#!/usr/bin/env python3
"""
Agent Geronimo v4 — Strategic Federal Opportunity Discovery & Triage

An elite opportunity scout, capture strategist, and institutional fit analyst.
Finds the BEST opportunities, not the MOST opportunities.

Searches broadly, filters aggressively, scores multi-dimensionally, and
produces a sharply curated pipeline optimized for decision usefulness.

Usage:
    python geronimo.py              # full run
    python geronimo.py --fresh      # clear cache first
"""
import os
import sys
import json
import hashlib
import re
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import click
import requests
import yaml
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
from rich.table import Table
from rich.panel import Panel
from dotenv import load_dotenv

console = Console()
PROJECT = Path(__file__).parent
load_dotenv(PROJECT / ".env")

# ═══════════════════════════════════════════════════════════
# HTTP helper
# ═══════════════════════════════════════════════════════════
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
})
TIMEOUT = 20
_last_req = 0.0


def _get(url: str, params: dict = None, timeout: int = TIMEOUT):
    """Rate-limited GET with retry."""
    global _last_req
    elapsed = time.time() - _last_req
    if elapsed < 0.6:
        time.sleep(0.6 - elapsed)
    _last_req = time.time()
    try:
        r = SESSION.get(url, params=params, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        return None


def _post(url: str, json_body: dict = None, headers: dict = None, timeout: int = TIMEOUT):
    """Rate-limited POST."""
    global _last_req
    elapsed = time.time() - _last_req
    if elapsed < 0.6:
        time.sleep(0.6 - elapsed)
    _last_req = time.time()
    try:
        r = SESSION.post(url, json=json_body, headers=headers, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception:
        return None


def _soup(url: str) -> Optional[BeautifulSoup]:
    """GET and parse HTML."""
    r = _get(url)
    if r:
        return BeautifulSoup(r.text, "lxml")
    return None


def _uid(source: str, text: str) -> str:
    return f"{source}_{hashlib.md5(text.encode()).hexdigest()[:12]}"


def verify_url(url: str) -> bool:
    """Verify a URL is reachable (returns 200-399). Uses HEAD with GET fallback."""
    if not url or not url.startswith("http"):
        return False
    try:
        r = SESSION.head(url, timeout=10, allow_redirects=True)
        if r.status_code < 400:
            return True
        # Some servers reject HEAD — fall back to GET
        r = SESSION.get(url, timeout=10, allow_redirects=True, stream=True)
        r.close()
        return r.status_code < 400
    except Exception:
        return False


def verify_all_urls(rows: list) -> list:
    """Verify every URL in the results; drop rows with dead links."""
    console.print(f"\n[bold]Verifying {len(rows)} URLs...[/bold]")
    verified = []
    dead = 0
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        console=console,
    ) as progress:
        task = progress.add_task("Checking links...", total=len(rows))
        for row in rows:
            url = row.get("url", "")
            if not url or not url.startswith("http"):
                dead += 1
                progress.advance(task)
                continue
            if verify_url(url):
                verified.append(row)
            else:
                dead += 1
            progress.advance(task)
    console.print(f"  Verified: [green]{len(verified)}[/green] live, [red]{dead}[/red] dead/removed")
    return verified


# ═══════════════════════════════════════════════════════════
# KEYWORD CONFIG (from old scraper's proven tier system)
# ═══════════════════════════════════════════════════════════
TIER1 = [  # 30 pts — direct mission hits
    # TSM / info warfare
    "Taiwan", "cognitive warfare", "information operations", "disinformation",
    "narrative warfare", "OSINT", "open source intelligence", "PRC influence",
    "China information", "Indo-Pacific security", "cross-strait", "Taiwan Strait",
    "coercion", "economic coercion", "military coercion", "coercive diplomacy",
    "cognitive domain", "information warfare", "perception management",
    "propaganda", "malign influence", "strategic deception",
    # Defense / intel tech
    "autonomous systems", "hypersonic", "missile defense", "electronic warfare",
    "undersea warfare", "space security", "critical infrastructure protection",
    "biodefense", "chemical defense", "directed energy", "counter-UAS",
    "artificial intelligence defense", "AI military", "machine learning defense",
    "cyber operations", "offensive cyber", "defensive cyber",
    "signals intelligence", "SIGINT", "HUMINT", "GEOINT", "MASINT",
    # Security policy
    "homeland security", "security policy", "counterterrorism",
    "intelligence community", "defense policy", "net assessment",
    "force design", "force structure", "defense strategy",
    # CONTRA / Latin America
    "terrorism", "counternarcotics", "organized crime", "Latin America security",
    "cartel", "illicit finance", "transnational crime", "gang violence",
    "Central America", "narcoterrorism",
]
TIER2 = [  # 20 pts — strong relevance
    "China", "deterrence", "Japan", "wargame", "wargaming", "crisis simulation",
    "narrative detection", "influence operations", "misinformation",
    "AI early warning", "Taiwan security", "East Asia security", "PRC", "CCP",
    "psychological operations", "counter-disinformation", "compellence",
    "gray zone coercion", "Beijing", "Chinese Communist Party",
    "computational propaganda", "information manipulation", "strategic competition",
    "Indo-Pacific strategy", "Asia-Pacific security", "INDOPACOM",
    "united front", "sharp power", "state-sponsored disinformation",
    "foreign interference", "election interference", "South Korea", "Korea",
    "semiconductor", "chip war", "technology competition",
    # Defense tech / APL
    "radar systems", "signal processing", "command and control", "C4ISR",
    "weapons systems", "force protection", "systems engineering",
    "threat assessment", "sensor systems", "unmanned systems", "UAS", "UAV",
    "quantum computing", "quantum sensing", "quantum cryptography",
    "5G military", "spectrum warfare", "EW", "electronic attack",
    "ballistic missile", "cruise missile", "anti-ship", "torpedo",
    "sonar", "acoustic sensing", "undersea", "submarine",
    "space domain awareness", "satellite", "orbital", "launch vehicle",
    "ISR", "intelligence surveillance reconnaissance",
    "deep learning", "computer vision", "natural language processing",
    "predictive analytics", "anomaly detection", "threat detection",
    # Security / CSPS
    "border security", "emergency management", "civil-military relations",
    "intelligence reform", "nuclear security", "arms control",
    "nonproliferation", "WMD", "weapons of mass destruction",
    "counterintelligence", "insider threat", "security clearance",
    # CONTRA
    "Venezuela", "Colombia", "Mexico security", "drug trafficking",
    "human trafficking", "money laundering", "terrorist financing",
    "radicalization", "extremism", "insurgency", "paramilitary",
    # Naval / maritime
    "naval warfare", "sea power", "fleet architecture", "littoral",
    "amphibious", "mine warfare", "anti-submarine",
]
TIER3 = [  # 10 pts — broader fit
    "Indo-Pacific", "Southeast Asia", "intelligence analysis",
    "geopolitical forecasting", "alliance resilience", "information resilience",
    "gray zone", "great power competition", "democracy resilience",
    "social media analysis", "NLP security", "military AI", "AI forecasting",
    "maritime security", "supply chain security", "authoritarian",
    "hybrid warfare", "narrative analysis", "strategic communication",
    "public diplomacy", "information environment", "media manipulation",
    "China policy", "Taiwan policy", "Asia security", "Pacific deterrence",
    "economic statecraft", "decoupling", "cybersecurity", "national security",
    "defense", "security studies", "foreign policy", "policy research",
    "resilience", "emerging technology",
    # Tech / APL broader
    "applied physics", "robotics", "machine learning", "neural network",
    "climate security", "health security", "pandemic preparedness",
    "additive manufacturing", "advanced materials", "nanotechnology",
    "biotechnology", "synthetic biology", "dual-use technology",
    "cloud security", "zero trust", "network security", "encryption",
    "data fusion", "sensor fusion", "multi-domain operations",
    # Policy broader
    "public policy", "governance", "regulatory policy",
    "crisis management", "risk assessment", "strategic planning",
    "alliance management", "burden sharing", "extended deterrence",
    "nuclear posture", "arms race", "escalation management",
    # CONTRA broader
    "rule of law", "democratic governance", "anti-corruption",
    "conflict resolution", "peacebuilding", "stabilization",
    "fragile states", "migration security", "refugee",
    "gang", "illicit trafficking", "border enforcement",
]
PRIORITY_FUNDERS = [  # +20 bonus
    # DoD / IC
    "DARPA", "SOCOM", "IARPA", "NGA", "INDOPACOM",
    "Army Research Laboratory", "Army Research Office",
    "Office of Naval Research", "Office of the Secretary of Defense",
    "Defense Intelligence", "Air Force Research", "Special Operations",
    "Missile Defense Agency", "DTRA", "Defense Threat Reduction",
    "Space Force", "Space Development Agency",
    "Office of Net Assessment", "OUSD",
    "Defense Advanced Research",
    "Naval Research Laboratory", "NSWC", "NAWC",
    "Air Force Office of Scientific Research",
    # Think tanks
    "Smith Richardson", "Minerva", "Carnegie", "Luce", "NED",
    "CNAS", "Center for a New American Security",
    "CSBA", "Center for Strategic and Budgetary",
    "Center for Naval Analyses",
    "RAND Corporation",
    "Hudson Institute", "Heritage Foundation",
    # Civilian / DHS / State
    "Johns Hopkins", "Department of Homeland Security", "DHS",
    "State Department", "Bureau of International Narcotics",
    "USAID", "Inter-American Foundation", "DEA",
    "Department of Energy", "NNSA",
    "FBI", "Secret Service",
]

# Grants.gov search terms
GG_SEARCH_TERMS = [
    # TSM / Asia / info warfare
    "Taiwan", "China", "Indo-Pacific", "cognitive warfare", "disinformation",
    "information operations", "OSINT", "influence operations", "deterrence",
    "open source intelligence", "misinformation", "psychological operations",
    "East Asia", "wargaming", "gray zone", "great power competition",
    "coercion", "Beijing", "CCP", "propaganda", "information warfare",
    "malign influence", "sharp power", "foreign interference", "hybrid warfare",
    "strategic competition", "computational propaganda", "cross-strait",
    "perception management", "compellence", "national security research",
    "security studies", "defense analysis", "cybersecurity research",
    "democratic resilience", "foreign policy research", "intelligence analysis",
    "emerging technology security", "Japan security", "Korea security",
    "South China Sea", "technology policy", "narrative analysis",
    "geospatial intelligence", "strategic communications", "conflict analysis",
    # Defense tech / AI / cyber
    "autonomous systems", "hypersonic", "missile defense", "electronic warfare",
    "undersea warfare", "space security", "critical infrastructure",
    "biodefense", "sensor systems", "command and control",
    "artificial intelligence", "machine learning", "deep learning",
    "cyber defense", "network security", "quantum computing",
    "directed energy", "counter-UAS", "unmanned systems",
    "signals intelligence", "ISR", "surveillance reconnaissance",
    "advanced computing", "data analytics", "predictive analytics",
    "5G security", "spectrum", "radar", "sonar",
    "naval research", "sea power", "anti-submarine warfare",
    "space domain", "satellite", "ballistic missile defense",
    # CSPS / homeland / intel
    "homeland security", "counterterrorism", "nuclear security",
    "arms control", "nonproliferation", "emergency management",
    "border security", "intelligence reform", "counterintelligence",
    "weapons of mass destruction", "chemical biological",
    "insider threat", "force protection",
    # CONTRA / Latin America
    "counternarcotics", "organized crime", "Latin America",
    "transnational crime", "drug trafficking", "terrorism research",
    "anti-corruption", "rule of law", "human trafficking",
    "Central America", "Colombia", "Venezuela", "Mexico security",
    "illicit finance", "gang violence", "border enforcement",
]

# SAM.gov search terms
# SAM.gov search terms — kept to ~15 broad terms to stay under 1,000 req/day quota
# Each broad term catches multiple narrower topics in a single API call
SAM_SEARCH_TERMS = [
    "Indo-Pacific security",          # Taiwan, China, PRC, East Asia, INDOPACOM
    "information operations",          # cognitive warfare, disinformation, influence ops, propaganda
    "intelligence analysis",           # OSINT, GEOINT, SIGINT, ISR, strategic warning
    "defense research",                # BAAs, defense innovation, applied research
    "cybersecurity",                   # cyber operations, network security, critical infrastructure
    "deterrence",                      # strategic competition, gray zone, hybrid warfare, coercion
    "counterterrorism",                # homeland security, threat assessment
    "autonomous systems",              # unmanned, AI, machine learning, robotics
    "missile defense",                 # hypersonic, directed energy, space systems
    "electronic warfare",              # EW, spectrum, sensor, C4ISR
    "naval research",                  # undersea, maritime, anti-submarine
    "counternarcotics",                # transnational crime, Latin America, organized crime
    "wargaming",                       # simulation, tabletop, scenario planning, net assessment
    "nuclear security",                # nonproliferation, arms control, WMD
    "social science research defense", # Minerva-style, political warfare, strategic comms
]

# NSF search terms (more targeted to avoid noise)
NSF_SEARCH_TERMS = [
    "Taiwan security", "China military", "Indo-Pacific security", "disinformation",
    "information operations", "influence operations", "OSINT intelligence",
    "deterrence strategy", "East Asia security", "wargaming simulation",
    "strategic communication", "open source intelligence", "information warfare",
    "hybrid warfare", "strategic competition", "computational propaganda",
    "cross-strait", "PLA military",
    "autonomous systems defense", "missile defense technology",
    "counterterrorism research", "homeland security research",
    "transnational crime", "Latin America governance", "organized crime",
    "nuclear nonproliferation", "arms control verification",
    "cybersecurity defense", "artificial intelligence security",
    "electronic warfare", "hypersonic defense", "quantum cryptography",
    "unmanned systems military", "undersea detection",
    "intelligence analysis", "geospatial intelligence",
]

# Organization project lines (for matching column)
# Format: (Display Name, comma-separated matching keywords)
ORG_PROJECTS = [
    # ── TSM (Taiwan Security Monitor) ──
    ("TSM: ADIZ Monitoring", "adiz, incursion, pla, military activity, air defense, fighter jet, sortie"),
    ("TSM: PRC Press Monitor", "press conference, mfa, tao, spokesperson, narrative, propaganda, prc"),
    ("TSM: SENTINEL Early Warning", "early warning, cross-strait, sentinel, crisis indicator, escalation"),
    ("TSM: OSINT/GEOINT Analysis", "osint, geoint, satellite, imagery, open source intelligence, remote sensing"),
    ("TSM: Narrative Warfare", "disinformation, cognitive warfare, information warfare, influence operation, misinformation"),
    ("TSM: Strategic Reports", "taiwan, indo-pacific, security brief, strategic analysis, strait"),
    ("TSM: Wargaming", "wargame, tabletop, crisis simulation, scenario, exercise, deterrence"),
    ("TSM: Policy Briefings", "policy brief, decision support, policymaker, defense policy"),
    ("TSM: Technology & Security", "artificial intelligence, cyber, emerging tech, autonomous, drone, surveillance"),
    # ── APL (Johns Hopkins Applied Physics Lab) ──
    ("APL: Autonomous Systems", "autonomous, robotics, unmanned, uav, drone, swarm"),
    ("APL: Missile Defense", "missile defense, hypersonic, ballistic, interceptor, aegis"),
    ("APL: Undersea & Maritime", "undersea, submarine, torpedo, sonar, maritime, naval"),
    ("APL: Electronic Warfare", "electronic warfare, jamming, spectrum, radar, signal processing, ew"),
    ("APL: Space Systems", "space, satellite, orbit, launch, space security, space domain"),
    ("APL: Cyber Operations", "cybersecurity, cyber operations, network defense, malware, cyber threat"),
    ("APL: Biodefense & Health", "biodefense, biosecurity, pandemic, chemical defense, health security"),
    ("APL: C4ISR", "command and control, c4isr, sensor, intelligence collection, surveillance, reconnaissance"),
    ("APL: Critical Infrastructure", "critical infrastructure, power grid, scada, industrial control"),
    # ── CSPS (Center for Security Policy Studies, GMU Schar School) ──
    ("CSPS: Homeland Security", "homeland security, border security, tsa, customs, immigration enforcement"),
    ("CSPS: Counterterrorism", "counterterrorism, terrorism, terrorist, radicalization, extremism, deradicalization"),
    ("CSPS: Intelligence Studies", "intelligence community, intelligence reform, intelligence analysis, cia, nsa, dia"),
    ("CSPS: Nuclear Security", "nuclear, nonproliferation, arms control, wmd, weapons of mass destruction, iaea"),
    ("CSPS: Emergency Management", "emergency management, disaster, fema, crisis response, resilience"),
    ("CSPS: Defense Policy", "defense policy, pentagon, dod, military strategy, force structure, national defense"),
    ("CSPS: Cybersecurity Policy", "cybersecurity policy, cyber deterrence, cyber norms, data privacy"),
    # ── CONTRA (Terrorism Research & Latin America Center, GMU) ──
    ("CONTRA: Narcoterrorism", "narcoterrorism, counternarcotics, drug trafficking, narcotics, drug cartel"),
    ("CONTRA: Organized Crime", "organized crime, cartel, transnational crime, criminal network, illicit"),
    ("CONTRA: Latin America", "latin america, central america, south america, venezuela, colombia, mexico, brazil, guatemala, honduras, el salvador"),
    ("CONTRA: Human Trafficking", "human trafficking, forced labor, modern slavery, smuggling"),
    ("CONTRA: Illicit Finance", "money laundering, illicit finance, terrorist financing, sanctions evasion, hawala"),
    ("CONTRA: Extremism", "violent extremism, radicalization, insurgency, paramilitary, militia"),
    ("CONTRA: Rule of Law", "rule of law, anti-corruption, governance, judicial reform, stabilization"),
    # ── CNA (Center for Naval Analyses) ──
    ("CNA: Naval Warfare", "naval warfare, sea power, fleet, maritime, ship, carrier, destroyer, frigate"),
    ("CNA: Marine Corps", "marine corps, amphibious, expeditionary, littoral"),
    ("CNA: Force Analysis", "force analysis, operational analysis, campaign analysis, modeling simulation"),
    # ── CNAS (Center for a New American Security) ──
    ("CNAS: Indo-Pacific", "indo-pacific, asia, china, alliance, partner, ally"),
    ("CNAS: Defense Program", "defense program, force design, defense budget, readiness, modernization"),
    ("CNAS: Tech & Security", "technology, artificial intelligence, autonomous, cyber, quantum, innovation"),
    ("CNAS: Energy & Climate", "energy security, climate, arctic, resource competition"),
    # ── CSBA (Center for Strategic and Budgetary Assessments) ──
    ("CSBA: Force Structure", "force structure, force planning, defense budget, capability, procurement"),
    ("CSBA: Operational Concepts", "operational concept, access, anti-access, area denial, A2/AD, power projection"),
    ("CSBA: Strategic Competition", "strategic competition, great power, china military, russia military"),
    # ── General defense/intel tech ──
    ("Defense AI/ML", "artificial intelligence, machine learning, deep learning, neural network, computer vision, nlp"),
    ("Cyber & Network Ops", "cybersecurity, cyber operations, network defense, zero trust, encryption, malware"),
    ("Autonomous & Unmanned", "autonomous, unmanned, drone, uav, uas, swarm, robotic"),
    ("Space & Satellite", "space, satellite, orbit, launch, gps, space domain, cislunar"),
    ("Quantum Technology", "quantum, quantum computing, quantum sensing, quantum cryptography, qubit"),
    ("Hypersonics & Missiles", "hypersonic, ballistic, cruise missile, missile defense, interceptor, glide"),
    ("EW & Spectrum", "electronic warfare, spectrum, jamming, radar, sigint, elint, comint"),
    ("Undersea & Acoustic", "undersea, submarine, torpedo, sonar, acoustic, anti-submarine, asw"),
    ("Bio & Chem Defense", "biodefense, biosecurity, chemical defense, cbrn, pandemic, pathogen"),
    ("Nuclear Deterrence", "nuclear, strategic deterrence, icbm, slbm, triad, nuclear posture"),
]


# ═══════════════════════════════════════════════════════════
# HARD FILTERS — exclude clearly irrelevant opportunities
# ═══════════════════════════════════════════════════════════
HARD_EXCLUDE_PATTERNS = re.compile(
    r'\b(biomedical|clinical trial|patient care|nursing|pharmaceutical|'
    r'veterinary|dental|optometry|podiatry|radiology|oncology|'
    r'highway construction|road paving|building renovation|'
    r'plumbing|hvac|janitorial|custodial|landscaping|'
    r'school lunch|child nutrition|head start|'
    r'municipal water|sewage treatment|waste disposal|'
    r'social work licensure|marriage counseling)\b',
    re.IGNORECASE,
)

SMALL_BUSINESS_ONLY = re.compile(
    r'\b(small business set.?aside|8\(a\) sole source|hubzone|'
    r'sdvosb set.?aside|wosb set.?aside|small business only)\b',
    re.IGNORECASE,
)


# ═══════════════════════════════════════════════════════════
# MULTI-DIMENSIONAL SCORING
# ═══════════════════════════════════════════════════════════
def _kw_in_text(kw: str, text: str) -> bool:
    """Check if keyword is in text, using word boundaries for short keywords."""
    kw_lower = kw.lower()
    if len(kw_lower) <= 4:
        return bool(re.search(r'\b' + re.escape(kw_lower) + r'\b', text))
    return kw_lower in text


def score_mission_fit(text: str, matched: list) -> int:
    """Score 1-5: how relevant is this to the institute's priorities?"""
    score = 0
    for kw in TIER1:
        if _kw_in_text(kw, text):
            score += 3
            matched.append(kw)
    for kw in TIER2:
        if _kw_in_text(kw, text):
            score += 2
            matched.append(kw)
    for kw in TIER3:
        if _kw_in_text(kw, text):
            score += 1
            matched.append(kw)
    for fkw in PRIORITY_FUNDERS:
        if _kw_in_text(fkw, text):
            score += 2
            matched.append(f"[FUNDER]{fkw}")
    # Map raw to 1-5
    if score >= 15:
        return 5
    if score >= 10:
        return 4
    if score >= 6:
        return 3
    if score >= 3:
        return 2
    return 1


def score_eligibility_fit(text: str, opp_type: str, funder: str) -> int:
    """Score 1-5: can the institute realistically pursue this?"""
    score = 3  # default — probably eligible
    text_lower = text.lower()
    opp_lower = opp_type.lower()

    # Boost for types clearly open to universities/nonprofits/think tanks
    if any(t in opp_lower for t in ["grant", "cooperative agreement", "fellowship",
                                     "baa", "rfi", "sources sought"]):
        score += 1
    if any(kw in text_lower for kw in ["university", "higher education", "nonprofit",
                                        "research institution", "think tank", "policy center"]):
        score += 1

    # Downgrade for things likely requiring cleared facilities or FFRDC status
    if any(kw in text_lower for kw in ["top secret", "ts/sci", "scif required",
                                        "ffrdc only", "uarc only", "cleared facility"]):
        score -= 2
    if SMALL_BUSINESS_ONLY.search(text):
        score -= 2

    # Contracts are harder for non-traditional performers
    if opp_lower in ("contract", "solicitation") and "research" not in text_lower:
        score -= 1

    return max(1, min(5, score))


def score_feasibility(text: str, opp_type: str, amount: str) -> int:
    """Score 1-5: how feasible is this to actually win?"""
    score = 3
    text_lower = text.lower()

    # Higher feasibility for research/analysis/policy
    if any(kw in text_lower for kw in ["policy research", "social science", "analysis",
                                        "assessment", "workshop", "convening",
                                        "report", "white paper", "study"]):
        score += 1

    # Lower feasibility for hardware/engineering-heavy
    if any(kw in text_lower for kw in ["prototype", "hardware", "manufacturing",
                                        "test range", "flight test", "systems integration",
                                        "production", "fabrication"]):
        score -= 1

    # BAAs and RFIs are more accessible
    opp_lower = opp_type.lower()
    if any(t in opp_lower for t in ["baa", "rfi", "sources sought"]):
        score += 1

    # Large contracts are harder to win for newcomers
    try:
        amt_str = re.sub(r'[^\d]', '', str(amount))
        if amt_str and int(amt_str) > 10_000_000:
            score -= 1
    except (ValueError, TypeError):
        pass

    return max(1, min(5, score))


def score_strategic_value(text: str, amount: str, funder: str) -> int:
    """Score 1-5: funding amount, prestige, repeatability, positioning value."""
    score = 3
    text_lower = text.lower()
    funder_lower = funder.lower()

    # Priority funders = higher strategic value
    for fkw in PRIORITY_FUNDERS:
        if _kw_in_text(fkw, funder_lower) or _kw_in_text(fkw, text_lower):
            score += 1
            break

    # Good funding amounts
    try:
        amt_str = re.sub(r'[^\d]', '', str(amount))
        if amt_str:
            amt = int(amt_str)
            if amt >= 500_000:
                score += 1
            elif amt < 25_000:
                score -= 1
    except (ValueError, TypeError):
        pass

    # Recurring/center/institute = long-term value
    if any(kw in text_lower for kw in ["center", "institute", "consortium",
                                        "recurring", "annual", "multi-year"]):
        score += 1

    return max(1, min(5, score))


def compute_total_score(mission: int, eligibility: int, feasibility: int, strategic: int) -> float:
    """Weighted total: 0.35*mission + 0.25*eligibility + 0.20*feasibility + 0.20*strategic."""
    raw = 0.35 * mission + 0.25 * eligibility + 0.20 * feasibility + 0.20 * strategic
    return round(raw, 2)


def determine_persona(text: str, opp_type: str) -> str:
    """Assign best institutional persona(s)."""
    text_lower = text.lower()
    opp_lower = opp_type.lower()
    personas = []

    if any(kw in text_lower for kw in ["university", "academic", "faculty", "research center",
                                        "higher education", "campus"]):
        personas.append("University center")
    if any(kw in text_lower for kw in ["policy", "strategic", "analysis", "assessment",
                                        "governance", "diplomatic", "foreign affairs"]):
        personas.append("Think tank")
    if any(kw in text_lower for kw in ["osint", "geoint", "data", "analytics", "monitoring",
                                        "forecasting", "early warning", "decision support",
                                        "AI", "machine learning", "modeling", "simulation"]):
        personas.append("Applied analytics shop")
    if any(kw in text_lower for kw in ["prototype", "engineering", "systems", "hardware",
                                        "technology development", "technical"]):
        personas.append("Technical teaming partner")

    if not personas:
        if "grant" in opp_lower or "cooperative" in opp_lower:
            personas.append("University center")
        elif "contract" in opp_lower or "baa" in opp_lower:
            personas.append("Think tank")
        else:
            personas.append("Think tank")

    return "; ".join(personas[:2])


def determine_mode(mission: int, eligibility: int, feasibility: int, text: str, opp_type: str) -> str:
    """Assign recommended mode: Prime, Sub, Team, Track, Discard."""
    text_lower = text.lower()
    opp_lower = opp_type.lower()

    if eligibility <= 1:
        return "Discard"

    if mission <= 1:
        return "Discard"

    # Research/policy opportunities — can likely prime
    if eligibility >= 4 and feasibility >= 3 and mission >= 3:
        if any(t in opp_lower for t in ["grant", "cooperative", "fellowship"]):
            return "Prime directly"
        if any(t in opp_lower for t in ["baa", "rfi", "sources sought"]):
            return "Prime through university"

    # Technical or large contracts — team or sub
    if any(kw in text_lower for kw in ["prototype", "hardware", "systems integration",
                                        "manufacturing", "cleared facility"]):
        if mission >= 3:
            return "Team with technical partner"
        return "Track for future cycle"

    if feasibility <= 2 and mission >= 3:
        return "Join as subawardee"

    if mission >= 3 and eligibility >= 3:
        if "contract" in opp_lower:
            return "Prime through university"
        return "Prime directly"

    if mission >= 2:
        return "Track for future cycle"

    return "Discard"


def generate_why_fits(matched_kw: list, funder: str, mission: int, persona: str) -> str:
    """Generate specific, non-generic explanation of fit."""
    topics = [kw for kw in matched_kw if not kw.startswith("[FUNDER]")][:6]
    funder_matches = [kw.replace("[FUNDER]", "") for kw in matched_kw if kw.startswith("[FUNDER]")]

    parts = []

    # Strength mapping
    strength_map = {
        "Taiwan": "Indo-Pacific strategic analysis and Taiwan security monitoring",
        "cognitive warfare": "cognitive warfare and information environment analysis",
        "information operations": "influence monitoring and information operations research",
        "disinformation": "disinformation detection and counter-narrative analysis",
        "OSINT": "open-source intelligence collection and analytical workflows",
        "deterrence": "deterrence theory and strategic competition analysis",
        "wargaming": "scenario-based wargaming and crisis simulation",
        "cybersecurity": "cyber policy research and critical infrastructure analysis",
        "Indo-Pacific": "Indo-Pacific regional security expertise",
        "gray zone": "gray-zone competition and sub-threshold operations analysis",
        "defense": "defense policy research and strategic studies",
        "national security": "national security research and policy analysis",
        "intelligence": "intelligence studies and strategic warning analysis",
        "missile defense": "defense technology assessment and force structure analysis",
        "electronic warfare": "defense technology and electronic warfare policy",
        "autonomous": "autonomous systems policy and AI security research",
    }

    for topic in topics[:3]:
        for key, desc in strength_map.items():
            if key.lower() in topic.lower():
                parts.append(desc)
                break

    if not parts:
        parts.append("security studies research and policy analysis")

    # Build the explanation
    core = parts[0] if parts else "national security research"
    extras = ", ".join(parts[1:3]) if len(parts) > 1 else ""

    explanation = f"Aligns with the institute's comparative advantage in {core}"
    if extras:
        explanation += f", as well as {extras}"
    explanation += "."

    if funder_matches:
        explanation += f" Priority funder ({', '.join(funder_matches[:2])}) with established relevance to the mission."

    if "University center" in persona:
        explanation += " Appears suited to a university-research model."
    elif "Applied analytics" in persona:
        explanation += " Fits the applied analytics and monitoring capability."
    elif "Think tank" in persona:
        explanation += " Matches policy analysis and strategic assessment capacity."

    return explanation


def generate_why_might_fail(eligibility: int, feasibility: int, text: str, opp_type: str) -> str:
    """Generate specific, skeptical risk assessment."""
    text_lower = text.lower()
    opp_lower = opp_type.lower()
    risks = []

    if any(kw in text_lower for kw in ["prototype", "hardware", "manufacturing",
                                        "test range", "flight test"]):
        risks.append("Appears heavily oriented toward hardware prototyping beyond this institute's standalone capacity.")
    if any(kw in text_lower for kw in ["ts/sci", "top secret", "scif", "classified"]):
        risks.append("Likely favors performers with classified infrastructure or prior DoD technical past performance.")
    if SMALL_BUSINESS_ONLY.search(text):
        risks.append("Small-business eligibility restriction suggests this is only viable through a startup partner.")
    if "contract" in opp_lower and "research" not in text_lower:
        risks.append("Service-type contract language may favor established defense contractors.")
    if any(kw in text_lower for kw in ["past performance", "prior contract", "cage code"]):
        risks.append("Likely requires significant past performance documentation that newer centers may lack.")
    if feasibility <= 2:
        risks.append("High competition level and technical demands may reduce capture probability.")
    if eligibility <= 2:
        risks.append("Eligibility constraints may prevent direct participation; teaming likely required.")

    if not risks:
        if opp_lower in ("contract", "solicitation"):
            risks.append("Contract vehicle may favor incumbents or performers with established government relationships.")
        else:
            risks.append("Review full solicitation to confirm eligibility and evaluation criteria alignment.")

    return " ".join(risks[:2])


def generate_concept_angle(matched_kw: list, text: str) -> str:
    """Generate a short, credible concept angle."""
    text_lower = text.lower()
    topics = [kw.lower() for kw in matched_kw if not kw.startswith("[FUNDER]")]

    angles = {
        "taiwan": "OSINT-enabled monitoring of gray-zone coercion and military signaling in the Taiwan Strait",
        "disinformation": "AI-assisted detection of coordinated narrative warfare campaigns targeting democratic institutions",
        "cognitive warfare": "Mapping cognitive warfare vectors and building institutional resilience frameworks",
        "wargaming": "Scenario-based wargaming of crisis escalation and democratic response under information attack",
        "information operations": "Systematic analysis of state-directed information operations and counter-strategy development",
        "deterrence": "Strategic-warning analytics for Indo-Pacific crisis instability and deterrence signaling",
        "osint": "Open-source intelligence workflows for real-time strategic monitoring and decision support",
        "cybersecurity": "Cyber policy and critical infrastructure vulnerability assessment with policy recommendations",
        "indo-pacific": "Alliance resilience and strategic competition analysis across the Indo-Pacific theater",
        "gray zone": "Sub-threshold competition monitoring and gray-zone response framework development",
        "missile defense": "Defense technology assessment and force posture analysis for emerging threat environments",
        "autonomous": "Policy frameworks for autonomous systems integration and AI-enabled security analysis",
        "counterterrorism": "Threat assessment and counter-network analysis for terrorism and transnational security",
        "counternarcotics": "Data-driven analysis of illicit networks and transnational organized crime disruption",
        "maritime": "Maritime coercion and alliance response modeling in contested littoral environments",
        "nuclear": "Nuclear posture assessment and arms control verification in the great-power competition context",
        "supply chain": "Supply-chain and research-security vulnerability mapping for critical technologies",
        "ai": "Human-centered AI governance and AI-enabled analysis for national security decision-making",
    }

    for topic in topics[:5]:
        for key, angle in angles.items():
            if key in topic:
                return angle

    # Fallback based on text content
    if "alliance" in text_lower or "partner" in text_lower:
        return "Alliance resilience analysis and partner-capacity assessment for strategic competition"
    if "innovation" in text_lower or "technology" in text_lower:
        return "Emerging technology assessment and defense innovation policy analysis"
    return "Strategic studies research with policy-relevant analytical deliverables"


def generate_suggested_partners(text: str, opp_type: str, feasibility: int) -> str:
    """Suggest plausible partner types."""
    text_lower = text.lower()
    partners = []

    if any(kw in text_lower for kw in ["AI", "machine learning", "data", "analytics", "algorithm"]):
        partners.append("AI/ML contractor or university CS department")
    if any(kw in text_lower for kw in ["cyber", "network", "infrastructure"]):
        partners.append("Cybersecurity lab or university cyber center")
    if any(kw in text_lower for kw in ["prototype", "hardware", "systems", "engineering"]):
        partners.append("Applied physics lab or engineering school")
    if any(kw in text_lower for kw in ["wargaming", "simulation", "modeling"]):
        partners.append("Wargaming/design contractor or war college")
    if any(kw in text_lower for kw in ["maritime", "naval", "undersea"]):
        partners.append("Maritime analytics shop or naval research partner")
    if any(kw in text_lower for kw in ["survey", "public opinion", "polling"]):
        partners.append("Polling/survey research group")
    if any(kw in text_lower for kw in ["democracy", "governance", "resilience", "civil society"]):
        partners.append("Democracy/resilience nonprofit organization")
    if any(kw in text_lower for kw in ["area studies", "regional", "language"]):
        partners.append("Regional studies center or area-studies faculty")
    if feasibility <= 2:
        partners.append("Federally connected research center with past performance")

    return "; ".join(partners[:3]) if partners else "University PI team or policy research partner"


def match_org_projects(title: str, description: str, matched_kw: list) -> str:
    """Match opportunity to TSM, APL, CSPS, and CONTRA project lines."""
    text = f"{title} {description}".lower()
    kw_text = " ".join(matched_kw).lower()
    combined = text + " " + kw_text
    matches = []
    for proj_name, proj_desc in ORG_PROJECTS:
        words = [w.strip() for w in re.split(r"[,]", proj_desc.lower()) if len(w.strip()) >= 3]
        if any(_kw_in_text(w, combined) for w in words):
            matches.append(proj_name)
    return "; ".join(matches[:4]) if matches else ""


# ═══════════════════════════════════════════════════════════
# SCRAPERS
# ═══════════════════════════════════════════════════════════

def scrape_grants_gov() -> list:
    """Grants.gov API — proven endpoint from old scraper."""
    console.print("  [cyan]Grants.gov API[/cyan]...", end="")
    results = {}
    for term in GG_SEARCH_TERMS:
        r = _post(
            "https://apply07.grants.gov/grantsws/rest/opportunities/search",
            json_body={"keyword": term, "oppStatuses": "posted|forecasted", "rows": 100},
            headers={"Content-Type": "application/json"},
        )
        if not r:
            continue
        try:
            for opp in r.json().get("oppHits", []):
                oid = str(opp.get("id", ""))
                if oid and oid not in results:
                    close_raw = opp.get("closeDate", "")
                    open_raw = opp.get("openDate", "")
                    # Parse dates
                    deadline = ""
                    if close_raw:
                        for fmt in ["%m/%d/%Y", "%m%d%Y"]:
                            try:
                                deadline = datetime.strptime(close_raw, fmt).strftime("%Y-%m-%d")
                                break
                            except ValueError:
                                deadline = close_raw
                    # Get award amounts
                    ceiling = opp.get("awardCeiling")
                    floor = opp.get("awardFloor")
                    amount = ""
                    if ceiling:
                        try:
                            amount = f"Up to ${int(ceiling):,}"
                        except (ValueError, TypeError):
                            amount = str(ceiling)
                    elif floor:
                        try:
                            amount = f"From ${int(floor):,}"
                        except (ValueError, TypeError):
                            amount = str(floor)

                    results[oid] = {
                        "id": f"gg_{oid}",
                        "title": opp.get("title", ""),
                        "funder": opp.get("agency", "Federal"),
                        "description": opp.get("synopsis", "") or "",
                        "deadline": deadline,
                        "amount": amount,
                        "url": f"https://www.grants.gov/search-results-detail/{oid}",
                        "source": "Grants.gov",
                        "opp_type": "Grant",
                    }
        except Exception:
            pass
    console.print(f" [green]{len(results)}[/green]")
    return list(results.values())


def scrape_sam_gov() -> list:
    """SAM.gov Opportunities API.

    Note: Requires a registered API key from https://open.gsa.gov/api/get-opportunities-public-api/
    DEMO_KEY will return 404. If no valid key, returns empty.
    """
    console.print("  [cyan]SAM.gov API[/cyan]...", end="")
    api_key = os.environ.get("SAM_GOV_API_KEY", "DEMO_KEY")
    if api_key == "DEMO_KEY":
        console.print(" [yellow]skipped (needs registered API key)[/yellow]")
        return []
    results = {}
    posted_from = (datetime.now() - timedelta(days=365)).strftime("%m/%d/%Y")
    errors = 0
    throttled = False
    for term in SAM_SEARCH_TERMS:
        if throttled:
            break  # Don't waste remaining quota
        r = _get(
            "https://api.sam.gov/prod/opportunities/v2/search",
            params={
                "api_key": api_key,
                "keyword": term,
                "postedFrom": posted_from,
                "ptype": "o,p,k,r,s,g,i",
                "limit": 100,
            },
        )
        if not r:
            errors += 1
            if errors == 1:
                console.print(f"\n    [yellow]SAM.gov query '{term}' failed — checking if throttled...[/yellow]")
            continue
        # Check for rate limit / throttle response
        try:
            body = r.json()
            if body.get("code") == "900804" or "throttled" in str(body.get("message", "")).lower():
                next_time = body.get("nextAccessTime", "unknown")
                console.print(f"\n    [red]SAM.gov API quota exceeded — resets at {next_time}[/red]")
                console.print(f"    [dim]Run again after quota resets to get SAM.gov results[/dim]")
                throttled = True
                continue
        except Exception:
            pass
        try:
            data = r.json()
            opp_list = data.get("opportunitiesData", [])
            if not opp_list:
                opp_list = data.get("_embedded", {}).get("results", [])
            for opp in opp_list:
                nid = opp.get("noticeId") or opp.get("solicitationNumber", "")
                if not nid or nid in results:
                    continue

                title = opp.get("title", "")
                if not title:
                    continue

                # Filter out awarded/closed/archived
                notice_type = str(opp.get("type", "")).lower()
                if any(x in notice_type for x in ["award", "cancel", "archive"]):
                    continue
                title_lower = title.lower()
                if any(x in title_lower for x in [
                    "award notice", "intent to sole source", "j&a -", "j&a-",
                    "modification ", "task order award",
                ]):
                    if not any(k in title_lower for k in ["funding available", "subcontract", "teaming"]):
                        continue

                # Check if deadline has passed
                close_raw = opp.get("responseDeadLine") or opp.get("archiveDate", "")
                deadline = ""
                if close_raw:
                    try:
                        from dateutil import parser as dp
                        dl_date = dp.parse(close_raw)
                        deadline = dl_date.strftime("%Y-%m-%d")
                        if dl_date < datetime.now() and notice_type not in ("p", "r"):
                            continue  # expired, skip
                    except Exception:
                        deadline = close_raw

                tmap = {"o": "Solicitation", "p": "Pre-Solicitation", "k": "Combined Synopsis",
                        "r": "Sources Sought", "s": "Special Notice", "g": "Grant", "i": "Intent to Bundle"}
                otype = tmap.get(opp.get("type", ""), "Opportunity")

                subtier = opp.get("subtierAgency", {})
                if isinstance(subtier, dict):
                    subtier = subtier.get("name", "")
                funder = subtier or opp.get("department", "") or "U.S. Government"

                results[nid] = {
                    "id": f"sam_{nid}",
                    "title": title,
                    "funder": funder,
                    "description": opp.get("description", ""),
                    "deadline": deadline,
                    "amount": "See solicitation",
                    "url": opp.get("uiLink") or f"https://sam.gov/opp/{nid}/view",
                    "source": "SAM.gov",
                    "opp_type": otype,
                }
        except Exception as e:
            if errors <= 3:
                console.print(f" [red]SAM parse error: {e}[/red]")
    console.print(f" [green]{len(results)}[/green]")
    return list(results.values())


def scrape_nsf_funding() -> list:
    """NSF open funding opportunities — actual solicitations you can apply to.

    Scrapes NSF's funding page for open Program Solicitations (PDs/PGAs),
    NOT past awards (which are already granted to other PIs).
    """
    console.print("  [cyan]NSF Funding Opportunities[/cyan]...", end="")
    results = {}

    # Search NSF's public funding opportunity listings via their search API
    nsf_search_terms = [
        "security", "defense", "intelligence", "cyber", "information",
        "disinformation", "Indo-Pacific", "Taiwan", "China", "deterrence",
        "autonomous", "missile", "electronic warfare", "AI", "machine learning",
        "quantum", "sensor", "undersea", "space", "satellite",
        "counterterrorism", "homeland", "nuclear", "arms control",
        "Latin America", "organized crime", "drug trafficking",
        "governance", "democracy", "conflict", "geospatial",
    ]

    for term in nsf_search_terms:
        r = _get(
            "https://www.nsf.gov/awardsearch/advancedSearchResult",
            params={
                "PIId": "",
                "PIFirstName": "",
                "PILastName": "",
                "PIOrganization": "",
                "PIState": "",
                "PIZip": "",
                "PICountry": "",
                "ProgOrganization": "",
                "ProgEleCode": "",
                "BooleanElement": "All",
                "ProgRefCode": "",
                "BooleanRef": "All",
                "Program": "",
                "ProgOfficer": "",
                "Keyword": term,
                "AwardNumberOperator": "",
                "AwardAmount": "",
                "AwardInstrument": "",
                "ActiveAwards": "true",
                "OriginalAwardDateOperator": "",
                "StartDateOperator": "",
                "ExpDateOperator": "",
            },
        )
        # We actually want open solicitations, not awards
        # The awards API gives us already-funded projects — skip those

    # Instead, scrape NSF's actual open funding opportunity pages
    nsf_urls = [
        "https://new.nsf.gov/funding/opportunities?sort_bef_combine=nsf_funding_upcoming_due_dates_702702_ASC",
        "https://www.nsf.gov/funding/pgm_list.jsp?org=SBE",  # Social/behavioral/economic sciences
        "https://www.nsf.gov/funding/pgm_list.jsp?org=CISE",  # Computer/information science
        "https://www.nsf.gov/funding/pgm_list.jsp?org=ENG",   # Engineering
    ]

    for url in nsf_urls:
        soup = _soup(url)
        if not soup:
            continue
        for a in soup.select("a[href]"):
            title = a.get_text(strip=True)
            if len(title) < 15 or len(title) > 200:
                continue
            # Look for actual program solicitations
            href = a.get("href", "")
            if not href:
                continue
            # NSF solicitation links contain /funding/ or pgm_summ
            if not any(p in href for p in ["/funding/", "pgm_summ", "solicitation", "pims_id"]):
                continue
            # Skip navigation and admin links
            if any(skip in title.lower() for skip in ["skip to", "sign in", "about", "contact", "faq"]):
                continue

            full_url = href if href.startswith("http") else f"https://www.nsf.gov{href}"

            # Check for security/defense relevance
            title_lower = title.lower()
            relevant = any(kw.lower() in title_lower for kw in [
                "security", "defense", "cyber", "intelligence", "information",
                "social", "behavioral", "decision", "human", "cognitive",
                "critical infrastructure", "network", "system", "data",
                "international", "political", "conflict", "governance",
            ])
            if not relevant:
                continue

            if title not in results:
                results[title] = {
                    "id": _uid("nsf_funding", title),
                    "title": title,
                    "funder": "NSF",
                    "description": "NSF open program solicitation — see listing for details and deadlines.",
                    "deadline": "See listing",
                    "amount": "Varies by program",
                    "url": full_url,
                    "source": "NSF Funding",
                    "opp_type": "Grant Solicitation",
                }

    console.print(f" [green]{len(results)}[/green]")
    return list(results.values())


def _web_opp(source, title, funder, desc, url, deadline="See website",
             amount="See website", opp_type="Funding Opportunity"):
    """Create a web-scraped opportunity dict."""
    return {
        "id": _uid(source, title + url),
        "title": title,
        "funder": funder,
        "description": desc,
        "deadline": deadline,
        "amount": amount,
        "url": url,
        "source": source,
        "opp_type": opp_type,
    }


def scrape_web_sources() -> list:
    """Scrape 20+ foundation and agency websites."""
    console.print("  [cyan]Web sources (30+ sites)[/cyan]...")
    all_results = []
    scrapers = [
        ("IARPA", _scrape_iarpa),
        ("DARPA", _scrape_darpa),
        ("ONR", _scrape_onr),
        ("Smith Richardson", _scrape_smith_richardson),
        ("Japan Foundation CGP", _scrape_japan_foundation),
        ("JSPS", _scrape_jsps),
        ("CCKF", _scrape_cckf),
        ("U.S.-Japan Foundation", _scrape_usjf),
        ("JUSFC", _scrape_jusfc),
        ("Taiwan Foundation for Democracy", _scrape_tfd),
        ("Wilson Center", _scrape_wilson),
        ("USIP", _scrape_usip),
        ("NED", _scrape_ned),
        ("Carnegie Corporation", _scrape_carnegie),
        ("MacArthur Foundation", _scrape_macarthur),
        ("Luce Foundation", _scrape_luce),
        ("Ploughshares Fund", _scrape_ploughshares),
        ("Open Society", _scrape_open_society),
        ("Atlantic Council", _scrape_atlantic_council),
        ("CSIS", _scrape_csis),
        ("Challenge.gov", _scrape_challenge_gov),
        ("Korea Foundation", _scrape_korea_foundation),
        # Defense / intel / think tanks
        ("RAND Corporation", _scrape_rand),
        ("CNAS", _scrape_cnas),
        ("CSBA", _scrape_csba),
        ("CNA", _scrape_cna),
        ("Hudson Institute", _scrape_hudson),
        ("Brookings", _scrape_brookings),
        ("AFOSR", _scrape_afosr),
        ("DHS S&T", _scrape_dhs_st),
        ("Stanton Foundation", _scrape_stanton),
        ("Inter-American Dialogue", _scrape_iad),
        ("Global Fund for Cyber", _scrape_gfce),
    ]
    for name, fn in scrapers:
        try:
            results = fn()
            all_results.extend(results)
            if results:
                console.print(f"    {name}: [green]{len(results)}[/green]")
        except Exception as e:
            console.print(f"    {name}: [red]error[/red]")
    return all_results


def _scrape_iarpa():
    results = []
    for url in ["https://www.iarpa.gov/research-programs/baa", "https://www.iarpa.gov/research-programs"]:
        soup = _soup(url)
        if not soup:
            continue
        for a in soup.select("a[href]"):
            title = a.get_text(strip=True)
            if len(title) < 10:
                continue
            href = a["href"]
            if any(kw in title.lower() or kw in href.lower() for kw in ["baa", "solicitation", "program", "research"]):
                full = href if href.startswith("http") else f"https://www.iarpa.gov{href}"
                results.append(_web_opp("IARPA", title, "IARPA",
                    "IARPA research program / BAA.", full, opp_type="BAA"))
    return results[:20]


def _scrape_darpa():
    results = []
    for url in ["https://www.darpa.mil/work-with-us/opportunities"]:
        soup = _soup(url)
        if not soup:
            continue
        for item in soup.select("article, .list-item, li"):
            t = item.select_one("h2, h3, h4, .title, a")
            if not t:
                continue
            title = t.get_text(strip=True)
            if len(title) < 10:
                continue
            link = t.get("href", "")
            if not link.startswith("http"):
                link = f"https://www.darpa.mil{link}" if link else url
            desc_el = item.select_one("p, .description")
            desc = desc_el.get_text(strip=True) if desc_el else "DARPA funding opportunity."
            results.append(_web_opp("DARPA", title, "DARPA", desc, link,
                amount="$1M-$10M+ typical", opp_type="BAA/Solicitation"))
    return results[:20]


def _scrape_onr():
    results = []
    soup = _soup("https://www.onr.navy.mil/work-with-us/funding-opportunities")
    if not soup:
        return []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) < 10:
            continue
        if any(kw in title.lower() for kw in ["baa", "broad agency", "solicitation", "announcement", "opportunity"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.onr.navy.mil{href}"
            results.append(_web_opp("ONR", title, "Office of Naval Research",
                "ONR research funding.", full, opp_type="BAA"))
    return results[:15]


def _scrape_smith_richardson():
    results = []
    for url in ["https://www.srf.org/programs/international-security-foreign-policy/", "https://www.srf.org/programs/"]:
        soup = _soup(url)
        if not soup:
            continue
        for sec in soup.select("article, .program, section"):
            t = sec.select_one("h1, h2, h3, h4")
            if not t:
                continue
            title = t.get_text(strip=True)
            if len(title) < 5:
                continue
            link = sec.select_one("a[href]")
            href = link["href"] if link else url
            if not href.startswith("http"):
                href = f"https://www.srf.org{href}"
            desc = sec.select_one("p")
            desc = desc.get_text(strip=True) if desc else "SRF international security & foreign policy grant."
            results.append(_web_opp("Smith Richardson", title, "Smith Richardson Foundation",
                desc, href, amount="$50K-$250K typical", opp_type="Foundation Grant"))
    return results[:10]


def _scrape_japan_foundation():
    results = []
    for url in ["https://www.jpf.go.jp/cgp/e/grant/", "https://www.cgp.org/grants"]:
        soup = _soup(url)
        if not soup:
            continue
        for a in soup.select("a[href]"):
            title = a.get_text(strip=True)
            if len(title) < 8 or not any(kw in title.lower() for kw in ["grant", "fellow", "program", "research", "exchange", "security"]):
                continue
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.jpf.go.jp{href}"
            results.append(_web_opp("Japan Foundation CGP", title, "Japan Foundation / CGP",
                "Japan-U.S. relations research grant.", full, amount="Varies", opp_type="Foundation Grant"))
    return results[:10]


def _scrape_jsps():
    soup = _soup("https://www.jsps.go.jp/english/e-fellow/index.html")
    if not soup:
        return []
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) < 10 or not any(kw in title.lower() for kw in ["research", "fellowship", "bilateral", "program"]):
            continue
        href = a["href"]
        full = href if href.startswith("http") else f"https://www.jsps.go.jp{href}"
        results.append(_web_opp("JSPS", title, "Japan Society for the Promotion of Science",
            "JSPS international research exchange.", full, amount="Varies", opp_type="Fellowship"))
    return results[:8]


def _scrape_cckf():
    soup = _soup("https://www.cckf.org/en/programs")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .program, li, .grant-item"):
        t = sec.select_one("h2, h3, h4, strong, b, a")
        if not t:
            continue
        title = t.get_text(strip=True)
        if len(title) < 8:
            continue
        link = sec.select_one("a[href]")
        href = link["href"] if link else "https://www.cckf.org/en/programs"
        if not href.startswith("http"):
            href = f"https://www.cckf.org{href}"
        results.append(_web_opp("CCKF", title, "Chiang Ching-kuo Foundation",
            "CCKF grant for Chinese studies / Taiwan-related research.", href,
            amount="$10K-$50K typical", opp_type="Foundation Grant"))
    return results[:10]


def _scrape_usjf():
    soup = _soup("https://us-jf.org/programs/")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .program, .grant"):
        t = sec.select_one("h2, h3, h4")
        if not t:
            continue
        title = t.get_text(strip=True)
        if len(title) < 5:
            continue
        link = sec.select_one("a[href]")
        href = link["href"] if link else "https://us-jf.org/programs/"
        if not href.startswith("http"):
            href = f"https://us-jf.org{href}"
        desc = sec.select_one("p")
        desc = desc.get_text(strip=True) if desc else "U.S.-Japan Foundation grant."
        results.append(_web_opp("USJF", title, "U.S.-Japan Foundation", desc, href,
            amount="Varies", opp_type="Foundation Grant"))
    return results[:8]


def _scrape_jusfc():
    soup = _soup("https://www.jusfc.gov/funding-opportunities/")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .entry, li"):
        t = sec.select_one("h2, h3, h4, a")
        if not t:
            continue
        title = t.get_text(strip=True)
        if len(title) < 8:
            continue
        link = sec.select_one("a[href]")
        href = link["href"] if link else "https://www.jusfc.gov/funding-opportunities/"
        if not href.startswith("http"):
            href = f"https://www.jusfc.gov{href}"
        results.append(_web_opp("JUSFC", title, "Japan-U.S. Friendship Commission",
            "JUSFC grant for U.S.-Japan research exchange.", href,
            amount="Varies", opp_type="Foundation Grant"))
    return results[:8]


def _scrape_tfd():
    results = []
    for url in ["https://www.tfd.org.tw/en/grants", "https://www.tfd.org.tw/en/programs"]:
        soup = _soup(url)
        if not soup:
            continue
        for a in soup.select("a[href]"):
            title = a.get_text(strip=True)
            if len(title) < 8 or not any(kw in title.lower() for kw in ["grant", "fellowship", "research", "democracy", "program"]):
                continue
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.tfd.org.tw{href}"
            results.append(_web_opp("TFD", title, "Taiwan Foundation for Democracy",
                "TFD grant supporting democracy research.", full, amount="Varies", opp_type="Foundation Grant"))
    return results[:10]


def _scrape_wilson():
    soup = _soup("https://www.wilsoncenter.org/fellowship-and-grant-opportunities")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .opportunity, .fellowship, .views-row"):
        t = sec.select_one("h2, h3, h4, .title, a")
        if not t:
            continue
        title = t.get_text(strip=True)
        if len(title) < 5:
            continue
        link = sec.select_one("a[href]")
        href = link["href"] if link else "https://www.wilsoncenter.org"
        if not href.startswith("http"):
            href = f"https://www.wilsoncenter.org{href}"
        desc = sec.select_one("p, .description")
        desc = desc.get_text(strip=True) if desc else "Wilson Center fellowship/grant."
        results.append(_web_opp("Wilson Center", title, "Wilson Center", desc, href,
            amount="Varies", opp_type="Fellowship/Grant"))
    return results[:8]


def _scrape_usip():
    soup = _soup("https://www.usip.org/grants-fellowships")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .grant, .fellowship, .views-row"):
        t = sec.select_one("h2, h3, h4, .title, a")
        if not t:
            continue
        title = t.get_text(strip=True)
        if len(title) < 5:
            continue
        link = sec.select_one("a[href]")
        href = link["href"] if link else "https://www.usip.org/grants-fellowships"
        if not href.startswith("http"):
            href = f"https://www.usip.org{href}"
        desc = sec.select_one("p")
        desc = desc.get_text(strip=True) if desc else "USIP grant/fellowship."
        results.append(_web_opp("USIP", title, "U.S. Institute of Peace", desc, href,
            amount="Up to $100K typical", opp_type="Grant/Fellowship"))
    return results[:8]


def _scrape_ned():
    soup = _soup("https://www.ned.org/apply-for-grant/en/")
    if not soup:
        soup = _soup("https://www.ned.org/apply-for-grant/")
    if not soup:
        return [_web_opp("NED", "NED Democracy Grants Program", "National Endowment for Democracy",
            "NED provides grants to support freedom worldwide. Areas: democratic governance, "
            "civil society, independent media, human rights. Active in 100+ countries.",
            "https://www.ned.org/apply-for-grant/", amount="$50K-$200K typical", opp_type="Foundation Grant")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in ["grant", "apply", "program", "fund"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.ned.org{href}"
            results.append(_web_opp("NED", title, "National Endowment for Democracy",
                "NED democracy promotion grant.", full, amount="$50K-$200K typical", opp_type="Foundation Grant"))
    if not results:
        results.append(_web_opp("NED", "NED Democracy Grants Program", "National Endowment for Democracy",
            "NED grants supporting freedom, democratic governance, civil society, independent media.",
            "https://www.ned.org/apply-for-grant/", amount="$50K-$200K typical", opp_type="Foundation Grant"))
    return results[:5]


def _scrape_carnegie():
    soup = _soup("https://www.carnegie.org/grants/grants-database/")
    if not soup:
        return [_web_opp("Carnegie", "Carnegie International Peace & Security Program",
            "Carnegie Corporation of New York",
            "Funds research on nuclear risk, international peace, technology & democracy. "
            "Supports think tanks, universities, and policy organizations.",
            "https://www.carnegie.org/grants/", amount="$100K-$500K typical", opp_type="Foundation Grant")]
    results = []
    for sec in soup.select("article, .grant, .views-row"):
        t = sec.select_one("h2, h3, h4, .title, a")
        if t:
            results.append(_web_opp("Carnegie", t.get_text(strip=True),
                "Carnegie Corporation", "", "https://www.carnegie.org/grants/",
                amount="$100K-$500K typical", opp_type="Foundation Grant"))
    if not results:
        results.append(_web_opp("Carnegie", "Carnegie International Peace & Security Program",
            "Carnegie Corporation of New York",
            "International peace, nuclear risk, technology & democracy research.",
            "https://www.carnegie.org/grants/", amount="$100K-$500K typical", opp_type="Foundation Grant"))
    return results[:5]


def _scrape_macarthur():
    return [_web_opp("MacArthur", "MacArthur Foundation - Nuclear Challenges & Big Bets",
        "MacArthur Foundation",
        "Funds nuclear risk reduction, climate, and bold systemic change initiatives. "
        "Big Bets program for transformative ideas. Supports universities and nonprofits.",
        "https://www.macfound.org/programs/", amount="$100K-$1M+", opp_type="Foundation Grant")]


def _scrape_luce():
    soup = _soup("https://www.hluce.org/programs/asia/")
    if not soup:
        return [_web_opp("Luce", "Luce Foundation Asia Program",
            "Henry Luce Foundation",
            "Deepening U.S. understanding of Asia through policy research, higher education, "
            "and public engagement. Strong focus on U.S.-Asia relations.",
            "https://www.hluce.org/programs/asia/", amount="$50K-$300K", opp_type="Foundation Grant")]
    results = []
    for sec in soup.select("article, section, .program"):
        t = sec.select_one("h2, h3, h4")
        if t and len(t.get_text(strip=True)) > 5:
            desc = sec.select_one("p")
            results.append(_web_opp("Luce", t.get_text(strip=True), "Henry Luce Foundation",
                desc.get_text(strip=True) if desc else "Luce Asia program grant.",
                "https://www.hluce.org/programs/asia/", amount="$50K-$300K", opp_type="Foundation Grant"))
    if not results:
        results.append(_web_opp("Luce", "Luce Foundation Asia Program", "Henry Luce Foundation",
            "U.S.-Asia understanding through policy research and education.",
            "https://www.hluce.org/programs/asia/", amount="$50K-$300K", opp_type="Foundation Grant"))
    return results[:5]


def _scrape_ploughshares():
    return [_web_opp("Ploughshares", "Ploughshares Fund - Nuclear Security Grants",
        "Ploughshares Fund",
        "Nuclear security, arms control, nonproliferation. Values policy advocacy and "
        "public education on nuclear threats. Active grantmaking.",
        "https://ploughshares.org/what-we-fund", amount="$25K-$150K", opp_type="Foundation Grant")]


def _scrape_open_society():
    return [_web_opp("Open Society", "Open Society Foundations - Grants",
        "Open Society Foundations",
        "Supports democracy, human rights, justice, governance. Funds media integrity, "
        "digital rights, government accountability programs worldwide.",
        "https://www.opensocietyfoundations.org/grants", amount="Varies widely", opp_type="Foundation Grant")]


def _scrape_atlantic_council():
    soup = _soup("https://www.atlanticcouncil.org/about/opportunities/")
    if not soup:
        return [_web_opp("Atlantic Council", "Atlantic Council Fellowships & Programs",
            "Atlantic Council",
            "Foreign policy, security, technology fellowships. Indo-Pacific Security Initiative, "
            "Digital Forensic Research Lab, Scowcroft Center programs.",
            "https://www.atlanticcouncil.org/about/opportunities/", amount="Varies", opp_type="Fellowship")]
    results = []
    for sec in soup.select("article, .views-row, li"):
        t = sec.select_one("h2, h3, h4, a")
        if t and len(t.get_text(strip=True)) > 8:
            results.append(_web_opp("Atlantic Council", t.get_text(strip=True),
                "Atlantic Council", "Atlantic Council program/fellowship.",
                "https://www.atlanticcouncil.org/about/opportunities/", amount="Varies", opp_type="Fellowship"))
    if not results:
        results.append(_web_opp("Atlantic Council", "Atlantic Council Fellowships & Programs",
            "Atlantic Council", "Foreign policy, security, technology fellowships.",
            "https://www.atlanticcouncil.org/about/opportunities/", amount="Varies", opp_type="Fellowship"))
    return results[:5]


def _scrape_csis():
    return [_web_opp("CSIS", "CSIS Research Fellowships & Programs",
        "Center for Strategic and International Studies",
        "Security studies, technology policy, Indo-Pacific program. Internships, "
        "fellowships, and commissioned research opportunities.",
        "https://www.csis.org/programs", amount="Varies", opp_type="Fellowship/Research")]


def _scrape_challenge_gov():
    soup = _soup("https://www.challenge.gov/?state=open")
    if not soup:
        return []
    results = []
    for sec in soup.select("article, .challenge-card, .views-row, li"):
        t = sec.select_one("h2, h3, h4, a, .title")
        if t and len(t.get_text(strip=True)) > 10:
            link = sec.select_one("a[href]")
            href = link["href"] if link else "https://www.challenge.gov"
            if not href.startswith("http"):
                href = f"https://www.challenge.gov{href}"
            results.append(_web_opp("Challenge.gov", t.get_text(strip=True),
                "Federal (Challenge.gov)", "Federal prize/challenge competition.",
                href, opp_type="Prize/Challenge"))
    return results[:10]


def _scrape_korea_foundation():
    soup = _soup("https://en.kf.or.kr/?menuno=3769")
    if not soup:
        return [_web_opp("Korea Foundation", "Korea Foundation Fellowship & Grant Programs",
            "Korea Foundation",
            "Supports Korea studies, policy research, cultural exchange. Fellowships for "
            "scholars, grants for Korean studies programs at universities.",
            "https://en.kf.or.kr/", amount="Varies", opp_type="Foundation Grant")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in ["grant", "fellowship", "program", "support"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://en.kf.or.kr{href}"
            results.append(_web_opp("Korea Foundation", title, "Korea Foundation",
                "Korea Foundation grant/fellowship.", full, amount="Varies", opp_type="Foundation Grant"))
    if not results:
        results.append(_web_opp("Korea Foundation", "Korea Foundation Fellowship & Grant Programs",
            "Korea Foundation", "Korea studies, policy research, cultural exchange.",
            "https://en.kf.or.kr/", amount="Varies", opp_type="Foundation Grant"))
    return results[:5]


def _scrape_rand():
    soup = _soup("https://www.rand.org/about/divisions.html")
    if not soup:
        return [_web_opp("RAND", "RAND Corporation Research Programs",
            "RAND Corporation",
            "Defense, homeland security, international affairs, national security, "
            "terrorism, Latin America, technology policy research and analysis.",
            "https://www.rand.org/about/divisions.html", amount="Varies", opp_type="Research")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in ["security", "defense", "terror", "latin", "homeland"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.rand.org{href}"
            results.append(_web_opp("RAND", title, "RAND Corporation",
                "RAND research program.", full, amount="Varies", opp_type="Research"))
    if not results:
        results.append(_web_opp("RAND", "RAND Corporation Research Programs",
            "RAND Corporation", "Defense, security, terrorism, policy research.",
            "https://www.rand.org/about/divisions.html", amount="Varies", opp_type="Research"))
    return results[:5]


def _scrape_iad():
    return [_web_opp("Inter-American Dialogue", "Inter-American Dialogue Programs & Fellowships",
        "Inter-American Dialogue",
        "Latin America policy research, democratic governance, migration, trade, "
        "rule of law, anti-corruption, security cooperation in the Americas.",
        "https://www.thedialogue.org/programs/", amount="Varies", opp_type="Fellowship/Research")]


def _scrape_brookings():
    soup = _soup("https://www.brookings.edu/careers/")
    if not soup:
        return [_web_opp("Brookings", "Brookings Institution Fellowships & Research",
            "Brookings Institution",
            "Foreign policy, governance, defense, global development, cybersecurity, "
            "Latin America, Middle East, East Asia research fellowships.",
            "https://www.brookings.edu/careers/", amount="Varies", opp_type="Fellowship/Research")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in ["fellow", "research", "scholar", "intern"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.brookings.edu{href}"
            results.append(_web_opp("Brookings", title, "Brookings Institution",
                "Brookings research fellowship/program.", full, amount="Varies", opp_type="Fellowship/Research"))
    if not results:
        results.append(_web_opp("Brookings", "Brookings Institution Fellowships & Research",
            "Brookings Institution", "Foreign policy, governance, defense research.",
            "https://www.brookings.edu/careers/", amount="Varies", opp_type="Fellowship/Research"))
    return results[:5]


def _scrape_cnas():
    soup = _soup("https://www.cnas.org/research")
    if not soup:
        return [_web_opp("CNAS", "CNAS Research Programs & Fellowships",
            "Center for a New American Security",
            "Indo-Pacific, defense, technology & national security, energy/climate, "
            "Middle East, transatlantic security. Fellowships for emerging leaders.",
            "https://www.cnas.org/research", amount="Varies", opp_type="Fellowship/Research")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in
            ["defense", "indo", "tech", "security", "fellow", "energy", "ai", "cyber", "china"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.cnas.org{href}"
            results.append(_web_opp("CNAS", title, "Center for a New American Security",
                "CNAS research program.", full, amount="Varies", opp_type="Research"))
    if not results:
        results.append(_web_opp("CNAS", "CNAS Research Programs & Fellowships",
            "Center for a New American Security",
            "Defense, technology, Indo-Pacific, energy/climate security research.",
            "https://www.cnas.org/research", amount="Varies", opp_type="Fellowship/Research"))
    return results[:10]


def _scrape_csba():
    return [
        _web_opp("CSBA", "CSBA Defense Strategy & Force Planning Research",
            "Center for Strategic and Budgetary Assessments",
            "Defense strategy, force structure analysis, operational concepts, A2/AD, "
            "power projection, great power competition, defense budget analysis.",
            "https://csbaonline.org/research", amount="Varies", opp_type="Research"),
        _web_opp("CSBA", "CSBA Commissioned Studies Program",
            "Center for Strategic and Budgetary Assessments",
            "Commissioned research on defense acquisition, operational concepts, "
            "net assessment, force design, and military technology trends.",
            "https://csbaonline.org/about/opportunities", amount="Varies", opp_type="Research"),
    ]


def _scrape_cna():
    soup = _soup("https://www.cna.org/careers")
    if not soup:
        return [_web_opp("CNA", "CNA Research Analyst & Fellowship Programs",
            "Center for Naval Analyses",
            "Naval warfare, Marine Corps operations, force analysis, homeland security, "
            "crisis management. Field representatives embedded with military commands.",
            "https://www.cna.org/careers", amount="$70K-$150K", opp_type="Research/Fellowship")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in
            ["research", "analyst", "fellow", "naval", "defense", "security"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.cna.org{href}"
            results.append(_web_opp("CNA", title, "Center for Naval Analyses",
                "CNA research opportunity.", full, amount="Varies", opp_type="Research"))
    if not results:
        results.append(_web_opp("CNA", "CNA Research Analyst & Fellowship Programs",
            "Center for Naval Analyses",
            "Naval warfare, force analysis, homeland security research.",
            "https://www.cna.org/careers", amount="$70K-$150K", opp_type="Research/Fellowship"))
    return results[:5]


def _scrape_hudson():
    soup = _soup("https://www.hudson.org/about/careers-internships")
    if not soup:
        return [_web_opp("Hudson", "Hudson Institute Fellowships & Research",
            "Hudson Institute",
            "National security, defense, foreign policy, technology. Known for "
            "Indo-Pacific, China, defense innovation, nuclear deterrence research.",
            "https://www.hudson.org/about/careers-internships", amount="Varies", opp_type="Fellowship/Research")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in
            ["fellow", "research", "scholar", "analyst", "defense", "security"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.hudson.org{href}"
            results.append(_web_opp("Hudson", title, "Hudson Institute",
                "Hudson Institute research program.", full, amount="Varies", opp_type="Fellowship/Research"))
    if not results:
        results.append(_web_opp("Hudson", "Hudson Institute Fellowships & Research",
            "Hudson Institute", "National security, defense, foreign policy research.",
            "https://www.hudson.org/about/careers-internships", amount="Varies", opp_type="Fellowship/Research"))
    return results[:5]


def _scrape_afosr():
    """Air Force Office of Scientific Research BAAs."""
    return [
        _web_opp("AFOSR", "AFOSR Broad Agency Announcement (BAA)",
            "Air Force Office of Scientific Research",
            "Basic research in physical sciences, engineering, life sciences, "
            "information & networks, human performance. Quantum, AI/ML, autonomy, "
            "directed energy, space, materials, cybersecurity.",
            "https://www.afrl.af.mil/AFOSR/", amount="$50K-$500K typical", opp_type="BAA"),
        _web_opp("AFOSR", "AFRL/AFOSR Young Investigator Research Program",
            "Air Force Office of Scientific Research",
            "Early-career scientists/engineers in defense-relevant basic research.",
            "https://www.afrl.af.mil/AFOSR/", amount="Up to $450K over 3 years", opp_type="Research Grant"),
    ]


def _scrape_dhs_st():
    """DHS Science & Technology Directorate."""
    soup = _soup("https://www.dhs.gov/science-and-technology/grants-and-funding")
    if not soup:
        return [_web_opp("DHS S&T", "DHS Science & Technology BAAs and Grants",
            "Department of Homeland Security S&T",
            "Cybersecurity, border security, counterterrorism technology, "
            "first responder tech, critical infrastructure protection, biometrics.",
            "https://www.dhs.gov/science-and-technology/grants-and-funding",
            amount="Varies ($100K-$5M)", opp_type="BAA/Grant")]
    results = []
    for a in soup.select("a[href]"):
        title = a.get_text(strip=True)
        if len(title) > 10 and any(kw in title.lower() for kw in
            ["grant", "baa", "funding", "solicitation", "research", "security", "cyber"]):
            href = a["href"]
            full = href if href.startswith("http") else f"https://www.dhs.gov{href}"
            results.append(_web_opp("DHS S&T", title, "Department of Homeland Security S&T",
                "DHS S&T research/grant opportunity.", full, amount="Varies", opp_type="BAA/Grant"))
    if not results:
        results.append(_web_opp("DHS S&T", "DHS Science & Technology BAAs and Grants",
            "Department of Homeland Security S&T",
            "Cybersecurity, border security, counterterrorism technology.",
            "https://www.dhs.gov/science-and-technology/grants-and-funding",
            amount="Varies ($100K-$5M)", opp_type="BAA/Grant"))
    return results[:5]


def _scrape_stanton():
    return [_web_opp("Stanton", "Stanton Foundation Nuclear Security Fellowships",
        "Stanton Foundation",
        "Nuclear security, arms control, nonproliferation fellowships. "
        "Supports early and mid-career scholars at major policy institutions.",
        "https://thestantonfoundation.org/", amount="$100K-$200K", opp_type="Fellowship")]


def _scrape_gfce():
    return [_web_opp("GFCE", "Global Forum on Cyber Expertise Programs",
        "Global Forum on Cyber Expertise",
        "International cyber capacity building, cyber norms, cyber diplomacy, "
        "national cybersecurity strategy development.",
        "https://thegfce.org/", amount="Varies", opp_type="Research/Capacity Building")]


# ═══════════════════════════════════════════════════════════
# LOCAL CONTEXT — scan TSM folder for project matching
# ═══════════════════════════════════════════════════════════
def load_local_context() -> set:
    """Read TSM!!! folder to extract project keywords."""
    tsm_path = Path("~/Desktop/TSM!!!").expanduser()
    keywords = set()
    if not tsm_path.exists():
        return keywords
    try:
        for item in tsm_path.rglob("*"):
            if item.is_file() and not item.name.startswith("~$"):
                name = item.stem.replace("_", " ").replace("-", " ").lower()
                for kw in ["taiwan", "tsm", "osint", "geoint", "adiz", "sentinel",
                           "narrative", "disinformation", "press", "monitor",
                           "pathfinder", "wargaming", "donor", "pitch", "grant",
                           "civil-military", "mobility", "pla", "early warning"]:
                    if kw in name:
                        keywords.add(kw)
                # Read .md and .txt for extra keywords
                if item.suffix.lower() in [".md", ".txt"]:
                    try:
                        text = item.read_text(errors="ignore")[:3000].lower()
                        for kw in ["taiwan", "osint", "narrative warfare", "cognitive warfare",
                                   "deterrence", "indo-pacific", "wargaming", "sentinel",
                                   "early warning", "adiz", "pla"]:
                            if kw in text:
                                keywords.add(kw)
                    except Exception:
                        pass
    except Exception:
        pass
    return keywords


# ═══════════════════════════════════════════════════════════
# DEDUPLICATION
# ═══════════════════════════════════════════════════════════
def deduplicate(opps: list) -> list:
    """Deduplicate by normalized title + funder."""
    seen = {}
    for opp in opps:
        key = re.sub(r"[^a-z0-9]", "", opp["title"].lower())[:60] + "|" + re.sub(r"[^a-z0-9]", "", opp["funder"].lower())[:20]
        if key not in seen or len(opp.get("description", "")) > len(seen[key].get("description", "")):
            seen[key] = opp
    return list(seen.values())


# ═══════════════════════════════════════════════════════════
# EXCEL EXPORT — clean columns, no empties
# ═══════════════════════════════════════════════════════════
EXCEL_COLUMNS = [
    ("title", "Opportunity Title", 48),
    ("funder", "Agency / Funder", 26),
    ("sub_agency", "Office / Program", 22),
    ("opp_type", "Type", 16),
    ("status", "Status", 12),
    ("total_score", "Total Score", 10),
    ("mission_fit", "Mission Fit (1-5)", 10),
    ("eligibility_fit", "Eligibility (1-5)", 10),
    ("feasibility", "Feasibility (1-5)", 10),
    ("strategic_value", "Strategic Value (1-5)", 10),
    ("persona", "Best Persona", 22),
    ("mode", "Recommended Mode", 22),
    ("why_fits", "Why It Fits", 55),
    ("why_might_fail", "Why It Might Fail", 45),
    ("concept_angle", "Concept Note Angle", 50),
    ("suggested_partners", "Suggested Partners", 35),
    ("project_match", "Relevant Programs", 35),
    ("synopsis", "Synopsis", 50),
    ("amount", "Funding / Range", 18),
    ("deadline", "Deadline", 14),
    ("url", "Official Link", 40),
    ("source", "Source Portal", 14),
    ("keywords_matched", "Key Topics / Tags", 38),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_excel(rows: list, stats: dict, output_dir: Path) -> str:
    """Write the Excel workbook with clean, populated columns."""
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"Agent_Geronimo_Funding_Results_{datetime.now():%Y_%m_%d}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()

    # ── Sheet 1: Pursue Now (Prime directly or Prime through university) ──
    ws = wb.active
    ws.title = "Pursue Now"
    pursue_rows = [r for r in rows if r.get("mode", "").startswith("Prime")]
    _write_sheet(ws, pursue_rows)

    # ── Sheet 2: Pursue with Partner (Sub, Team) ──
    ws2 = wb.create_sheet("Pursue with Partner")
    partner_rows = [r for r in rows if r.get("mode", "") in ("Join as subawardee", "Team with technical partner")]
    _write_sheet(ws2, partner_rows)

    # ── Sheet 3: Monitor (Track for future cycle) ──
    ws3 = wb.create_sheet("Monitor")
    monitor_rows = [r for r in rows if r.get("mode", "") == "Track for future cycle"]
    _write_sheet(ws3, monitor_rows)

    # ── Sheet 4: All Opportunities ──
    ws4 = wb.create_sheet("All Opportunities")
    _write_sheet(ws4, rows)

    # ── Sheet 5: Federal (Grants.gov + SAM.gov) ──
    fed_rows = [r for r in rows if r["source"] in ("Grants.gov", "SAM.gov")]
    ws5 = wb.create_sheet("Federal")
    _write_sheet(ws5, fed_rows)

    # ── Sheet 6: Foundations & Think Tanks ──
    found_rows = [r for r in rows if "Foundation" in r.get("opp_type", "") or "Fellowship" in r.get("opp_type", "")
                  or r["source"] not in ("Grants.gov", "SAM.gov", "NSF Funding")]
    ws6 = wb.create_sheet("Foundations & Think Tanks")
    _write_sheet(ws6, found_rows)

    # ── Sheet 7: Executive Summary ──
    ws7 = wb.create_sheet("Executive Summary")
    ws7.column_dimensions["A"].width = 40
    ws7.column_dimensions["B"].width = 60
    _write_executive_summary(ws7, rows, stats)

    wb.save(filepath)
    return str(filepath)


def _write_sheet(ws, rows):
    """Write header + data rows to a worksheet."""
    for ci, (_, header, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 55)

    for ri, row in enumerate(rows, 2):
        for ci, (field, _, _) in enumerate(EXCEL_COLUMNS, 1):
            val = row.get(field, "")
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        # Hyperlink on URL column
        url_col = next(i for i, (f, _, _) in enumerate(EXCEL_COLUMNS, 1) if f == "url")
        url_val = row.get("url", "")
        if url_val and url_val.startswith("http"):
            cell = ws.cell(row=ri, column=url_col)
            try:
                cell.hyperlink = url_val
                cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass

        # Color by mode
        mode = row.get("mode", "")
        fill = None
        if mode.startswith("Prime"):
            fill = GREEN_FILL
        elif mode in ("Join as subawardee", "Team with technical partner"):
            fill = YELLOW_FILL
        elif mode == "Track for future cycle":
            fill = ORANGE_FILL
        if fill:
            for ci in range(1, len(EXCEL_COLUMNS) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _write_executive_summary(ws, rows: list, stats: dict):
    """Write executive summary sheet with top priorities and landscape analysis."""
    bold = Font(bold=True, size=11)
    header = Font(bold=True, size=13, color="1F4E79")

    r = 1
    ws.cell(row=r, column=1, value="AGENT GERONIMO v4 — EXECUTIVE SUMMARY").font = header
    r += 1
    ws.cell(row=r, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = Font(italic=True)
    r += 2

    # Pipeline overview
    ws.cell(row=r, column=1, value="Pipeline Overview").font = bold
    r += 1
    pursue_now = [x for x in rows if x.get("mode", "").startswith("Prime")]
    pursue_partner = [x for x in rows if x.get("mode", "") in ("Join as subawardee", "Team with technical partner")]
    monitor = [x for x in rows if x.get("mode", "") == "Track for future cycle"]
    overview = [
        ("Total raw opportunities scraped", stats.get("raw", 0)),
        ("After deduplication", stats.get("deduped", 0)),
        ("After hard filters + scoring", len(rows)),
        ("Pursue Now (Prime)", len(pursue_now)),
        ("Pursue with Partner (Sub/Team)", len(pursue_partner)),
        ("Monitor (Track)", len(monitor)),
        ("Grants.gov results", stats.get("grants_gov", 0)),
        ("SAM.gov results", stats.get("sam_gov", 0)),
        ("NSF Funding results", stats.get("nsf", 0)),
        ("Web source results", stats.get("web", 0)),
    ]
    for label, val in overview:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=str(val))
        r += 1
    r += 1

    # Top 10 priorities
    ws.cell(row=r, column=1, value="Top 10 Priorities").font = bold
    r += 1
    for i, row in enumerate(pursue_now[:10], 1):
        ws.cell(row=r, column=1, value=f"{i}. {row['title'][:80]}")
        ws.cell(row=r, column=2, value=f"{row['funder']} | {row.get('mode','')} | Score: {row.get('total_score','')}")
        r += 1
    r += 1

    # Top 5 partnership opportunities
    ws.cell(row=r, column=1, value="Top 5 Partnership Opportunities").font = bold
    r += 1
    for i, row in enumerate(pursue_partner[:5], 1):
        ws.cell(row=r, column=1, value=f"{i}. {row['title'][:80]}")
        ws.cell(row=r, column=2, value=f"{row['funder']} | Partners: {row.get('suggested_partners','')[:60]}")
        r += 1
    r += 1

    # Agencies to watch
    ws.cell(row=r, column=1, value="Top Funders by Volume").font = bold
    r += 1
    from collections import Counter
    funder_counts = Counter(row["funder"] for row in rows)
    for funder, count in funder_counts.most_common(10):
        ws.cell(row=r, column=1, value=funder)
        ws.cell(row=r, column=2, value=f"{count} opportunities")
        r += 1
    r += 1

    # Notes
    ws.cell(row=r, column=1, value="Notes").font = bold
    r += 1
    notes = [
        "Scores are multi-dimensional (Mission Fit, Eligibility, Feasibility, Strategic Value) — each 1-5.",
        "Total Score = 0.35*Mission + 0.25*Eligibility + 0.20*Feasibility + 0.20*Strategic.",
        "All URLs verified live — dead links removed before export.",
        "Hard filters excluded biomedical, construction, janitorial, and other clearly irrelevant categories.",
        "Small-business set-aside opportunities flagged — may require teaming with SB partner.",
        "Deadlines may change — always check the original listing.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        r += 1


# ═══════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════
@click.command()
@click.option("--fresh", is_flag=True, help="Clear cache before running")
def main(fresh: bool):
    console.print(Panel(
        "[bold cyan]AGENT GERONIMO v4[/bold cyan]\n"
        "[dim]Strategic Federal Opportunity Discovery & Triage[/dim]",
        border_style="cyan",
    ))

    # Stage 1: Discovery
    console.print("\n[bold]Stage 1: Discovery[/bold]")
    all_opps = []
    stats = {}

    gg = scrape_grants_gov()
    stats["grants_gov"] = len(gg)
    all_opps.extend(gg)

    sam = scrape_sam_gov()
    stats["sam_gov"] = len(sam)
    all_opps.extend(sam)

    nsf = scrape_nsf_funding()
    stats["nsf"] = len(nsf)
    all_opps.extend(nsf)

    web = scrape_web_sources()
    stats["web"] = len(web)
    all_opps.extend(web)

    stats["raw"] = len(all_opps)
    console.print(f"\n  Raw total: [bold]{stats['raw']}[/bold]")

    # Stage 2: Dedup
    console.print("\n[bold]Stage 2: Deduplication[/bold]")
    opps = deduplicate(all_opps)
    stats["deduped"] = len(opps)
    console.print(f"  {stats['raw']} → [green]{stats['deduped']}[/green] unique")

    # Stage 3: Hard Filter + Score + Enrich
    console.print("\n[bold]Stage 3: Filter, Score & Enrich[/bold]")
    local_kw = load_local_context()
    if local_kw:
        console.print(f"  Local TSM context: {len(local_kw)} keywords from TSM!!! folder")

    scored_rows = []
    filtered_out = 0
    today = datetime.now()
    for opp in opps:
        title = opp.get("title", "")
        desc = opp.get("description", "")
        funder = opp.get("funder", "")
        opp_type = opp.get("opp_type", "")
        amount = opp.get("amount", "")
        combined_text = f"{title} {desc}".lower()

        # Skip opportunities with deadlines that have already passed
        dl = opp.get("deadline", "")
        if dl and dl not in ("See listing", "See website", "See solicitation", "Ongoing", "TBD", "Rolling", ""):
            try:
                dl_date = datetime.strptime(dl, "%Y-%m-%d")
                if dl_date < today:
                    continue  # expired
            except ValueError:
                pass  # unparseable deadline — keep it

        # Hard filter: exclude clearly irrelevant opportunities
        if HARD_EXCLUDE_PATTERNS.search(combined_text):
            filtered_out += 1
            continue

        # Multi-dimensional scoring
        matched = []
        mission = score_mission_fit(combined_text, matched)
        eligibility = score_eligibility_fit(combined_text, opp_type, funder)
        feasibility = score_feasibility(combined_text, opp_type, amount)
        strategic = score_strategic_value(combined_text, amount, funder)
        total = compute_total_score(mission, eligibility, feasibility, strategic)

        # Skip very low relevance (mission 1 = no keyword hits at all)
        if mission <= 1 and not any(f.lower() in funder.lower() for f in PRIORITY_FUNDERS[:20]):
            filtered_out += 1
            continue

        # Classify
        persona = determine_persona(combined_text, opp_type)
        mode = determine_mode(mission, eligibility, feasibility, combined_text, opp_type)

        # Skip discards
        if mode == "Discard":
            filtered_out += 1
            continue

        # Generate analytical fields
        why_fits = generate_why_fits(matched, funder, mission, persona)
        why_might_fail = generate_why_might_fail(eligibility, feasibility, combined_text, opp_type)
        concept_angle = generate_concept_angle(matched, combined_text)
        suggested_partners = generate_suggested_partners(combined_text, opp_type, feasibility)
        proj_match = match_org_projects(title, desc, matched)

        scored_rows.append({
            "title": title,
            "funder": funder,
            "sub_agency": "",  # SAM.gov populates this; others left blank
            "opp_type": opp_type,
            "status": "Active",
            "total_score": total,
            "mission_fit": mission,
            "eligibility_fit": eligibility,
            "feasibility": feasibility,
            "strategic_value": strategic,
            "persona": persona,
            "mode": mode,
            "why_fits": why_fits,
            "why_might_fail": why_might_fail,
            "concept_angle": concept_angle,
            "suggested_partners": suggested_partners,
            "project_match": proj_match if proj_match else "General relevance — review for fit",
            "synopsis": desc[:500] if desc else "See listing",
            "amount": amount or "See listing",
            "deadline": dl or "See listing",
            "url": opp.get("url", ""),
            "source": opp.get("source", ""),
            "keywords_matched": ", ".join([k for k in matched if not k.startswith("[FUNDER]")][:8]),
        })

    # Sort by total score descending, then mission fit
    scored_rows.sort(key=lambda r: (r["total_score"], r["mission_fit"]), reverse=True)

    console.print(f"  Hard-filtered: [red]{filtered_out}[/red] irrelevant")
    console.print(f"  Scored & enriched: [green]{len(scored_rows)}[/green] opportunities")

    # Stage 3b: URL Verification — remove dead links
    console.print("\n[bold]Stage 3b: URL Verification[/bold]")
    scored_rows = verify_all_urls(scored_rows)

    pursue_count = sum(1 for r in scored_rows if r["mode"].startswith("Prime"))
    partner_count = sum(1 for r in scored_rows if r["mode"] in ("Join as subawardee", "Team with technical partner"))
    monitor_count = sum(1 for r in scored_rows if r["mode"] == "Track for future cycle")

    console.print(f"  Pursue Now: [green]{pursue_count}[/green]")
    console.print(f"  Pursue w/ Partner: [yellow]{partner_count}[/yellow]")
    console.print(f"  Monitor: [dim]{monitor_count}[/dim]")

    # Stage 4: Export
    console.print("\n[bold]Stage 4: Export[/bold]")
    output_dir = PROJECT / "output"
    excel_path = export_excel(scored_rows, stats, output_dir)
    console.print(f"  Excel: [green]{excel_path}[/green]")

    # Also copy to TSM folder
    tsm_out = Path("~/Desktop/TSM!!!/Agent Geronimo").expanduser()
    tsm_out.mkdir(parents=True, exist_ok=True)
    import shutil
    shutil.copy2(excel_path, tsm_out / Path(excel_path).name)
    console.print(f"  Copied to: [green]{tsm_out}[/green]")

    # CSV backup
    csv_path = output_dir / f"Agent_Geronimo_Results_{datetime.now():%Y_%m_%d}.csv"
    pd.DataFrame(scored_rows).to_csv(csv_path, index=False)
    console.print(f"  CSV: [green]{csv_path}[/green]")

    # Summary
    console.print(f"""
{'='*60}
  AGENT GERONIMO v4 — COMPLETE
{'='*60}
  Raw opportunities:    {stats['raw']}
  After dedup:          {stats['deduped']}
  After scoring:        {len(scored_rows)}
  Pursue Now:           {pursue_count}
  Pursue w/ Partner:    {partner_count}
  Monitor:              {monitor_count}
  Grants.gov:           {stats['grants_gov']}
  SAM.gov:              {stats['sam_gov']}
  NSF Funding:          {stats['nsf']}
  Web sources:          {stats['web']}
{'='*60}
""")

    # Top opportunities table
    table = Table(title="Top 20 Opportunities")
    table.add_column("Score", width=5, style="bold")
    table.add_column("Title", width=44)
    table.add_column("Funder", width=22)
    table.add_column("Mode", width=20)
    table.add_column("Mission", width=7)
    table.add_column("Source", width=12)
    for r in scored_rows[:20]:
        mode = r.get("mode", "")
        style = "green" if mode.startswith("Prime") else "yellow" if "sub" in mode.lower() or "team" in mode.lower() else "dim"
        table.add_row(
            str(r["total_score"]), r["title"][:44], r["funder"][:22],
            mode[:20], str(r["mission_fit"]), r["source"][:12], style=style,
        )
    console.print(table)


if __name__ == "__main__":
    main()
