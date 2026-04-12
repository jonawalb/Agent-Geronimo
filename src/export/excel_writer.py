"""Excel workbook export with multiple sheets, formatting, and hyperlinks.

Produces the primary Agent Geronimo output: a polished Excel workbook
with all opportunity data, scoring, recommendations, and source tracking.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import Opportunity, RunStats

logger = logging.getLogger("geronimo.export.excel")

# Column definitions for the Master sheet
MASTER_COLUMNS = [
    ("opportunity_id", "Opportunity ID", 15),
    ("title", "Opportunity Title", 45),
    ("opportunity_type", "Type", 12),
    ("funder", "Funder / Agency", 25),
    ("sub_agency", "Sub-Agency / Office", 20),
    ("listing_url", "Listing URL", 30),
    ("application_url", "Application Portal", 30),
    ("synopsis", "Synopsis", 50),
    ("eligibility_text", "Eligibility", 30),
    ("universities_eligible", "Universities?", 12),
    ("research_centers_eligible", "Research Centers?", 12),
    ("nonprofits_eligible", "Nonprofits?", 12),
    ("think_tanks_eligible", "Think Tanks?", 12),
    ("university_centers_eligible", "Univ Centers?", 12),
    ("geographic_restrictions", "Geo Restrictions", 15),
    ("citizenship_restrictions", "Citizenship/Security", 15),
    ("topic_area", "Topic Area", 20),
    ("keywords", "Keywords", 30),
    ("security_defense_relevance", "Security/Defense", 12),
    ("taiwan_relevance", "Taiwan", 10),
    ("indo_pacific_relevance", "Indo-Pacific", 12),
    ("info_warfare_relevance", "Info/Narrative War", 12),
    ("cyber_tech_relevance", "Cyber/Tech", 10),
    ("policy_center_relevance", "Policy Center", 12),
    ("tsm_fit_score", "TSM Fit", 8),
    ("gmu_center_fit_score", "GMU Fit", 8),
    ("general_security_fit_score", "General Fit", 8),
    ("overall_relevance_score", "Overall Score", 10),
    ("confidence_score", "Confidence", 10),
    ("estimated_competitiveness", "Competitiveness", 14),
    ("estimated_difficulty", "Difficulty", 12),
    ("deadline", "Deadline", 14),
    ("open_date", "Open Date", 14),
    ("recurring_cycle", "Recurring Cycle", 14),
    ("award_min", "Award Min", 14),
    ("award_max", "Award Max", 14),
    ("typical_award", "Typical Award", 14),
    ("cost_share", "Cost Share", 12),
    ("project_length", "Project Length", 12),
    ("num_awards_expected", "# Awards", 10),
    ("prior_awardees", "Prior Awardees", 30),
    ("prior_awardee_links", "Prior Awardee Links", 20),
    ("similar_projects", "Similar Projects", 30),
    ("similar_project_summaries", "Similar Summaries", 30),
    ("funder_preferences", "Funder Preferences", 40),
    ("suggested_framing_tsm", "TSM Framing", 40),
    ("suggested_framing_gmu", "GMU/CSPS Framing", 40),
    ("suggested_proposal_angle", "Proposal Angle", 40),
    ("suggested_concept_paragraph", "Concept Paragraph", 50),
    ("suggested_outline", "Outline", 30),
    ("recommended_lead_type", "Lead Type", 20),
    ("recommended_next_step", "Next Step", 30),
    ("urgency", "Urgency", 14),
    ("notes", "Notes", 30),
    ("red_flags", "Red Flags", 20),
    ("data_verified_date", "Verified Date", 12),
    ("scrape_date", "Scrape Date", 12),
    ("source_citations", "Sources", 20),
    ("tsm_pipeline_fit", "TSM Pipeline Fit", 20),
    ("tsm_pipeline_explanation", "Pipeline Explanation", 40),
    ("existing_concept_match", "Concept Match", 20),
    ("internal_file_references", "Internal Files", 20),
    ("final_recommendation", "Recommendation", 14),
    ("why_this_could_work", "Why This Could Work", 55),
    ("all_source_urls", "All Source URLs", 25),
]

# Color scheme
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HIGH_PRIORITY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MEDIUM_PRIORITY_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_PRIORITY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


class ExcelWriter:
    """Exports opportunities to a formatted Excel workbook."""

    def __init__(self, output_dir: str = "~/agent-geronimo/output"):
        self.output_dir = Path(output_dir).expanduser()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, opportunities: List[Opportunity], stats: RunStats) -> str:
        """Export all opportunities to a formatted Excel workbook.

        Returns the file path of the generated workbook.
        """
        filename = f"Agent_Geronimo_Funding_Results_{datetime.now():%Y_%m_%d}.xlsx"
        filepath = self.output_dir / filename

        wb = Workbook()

        # Sheet 1: Master Opportunities
        self._write_master_sheet(wb, opportunities)

        # Sheet 2: Top Priority - TSM
        tsm_top = [o for o in opportunities if o.tsm_fit_score >= 15]
        tsm_top.sort(key=lambda o: o.tsm_fit_score, reverse=True)
        self._write_filtered_sheet(wb, "Top Priority - TSM", tsm_top)

        # Sheet 3: Top Priority - GMU Center
        gmu_top = [o for o in opportunities if o.gmu_center_fit_score >= 15]
        gmu_top.sort(key=lambda o: o.gmu_center_fit_score, reverse=True)
        self._write_filtered_sheet(wb, "Top Priority - GMU Center", gmu_top)

        # Sheet 4: Federal Grants
        federal = [o for o in opportunities if o.opportunity_type == "Grant"
                   and any(kw in (o.funder or "").lower() for kw in
                          ["federal", "department", "dod", "dhs", "nsf", "doe",
                           "defense", "state", "army", "navy", "air force"])]
        self._write_filtered_sheet(wb, "Federal Grants", federal)

        # Sheet 5: Contracts / BAAs / RFPs
        contracts = [o for o in opportunities
                     if o.opportunity_type in ["Contract", "BAA", "RFP", "RFI",
                                                "Cooperative Agreement"]]
        self._write_filtered_sheet(wb, "Contracts BAAs RFPs", contracts)

        # Sheet 6: Foundations
        foundations = [o for o in opportunities
                      if any(kw in (o.funder or "").lower() for kw in
                            ["foundation", "endowment", "fund", "carnegie", "macarthur",
                             "ford", "hewlett", "luce", "smith richardson", "rockefeller",
                             "ned", "open society", "stanton", "ploughshares"])]
        self._write_filtered_sheet(wb, "Foundations", foundations)

        # Sheet 7: Past Award Analysis (placeholder with funder profiles)
        self._write_award_analysis_sheet(wb)

        # Sheet 8: Source Log
        self._write_source_log(wb, stats)

        # Sheet 9: Run Notes
        self._write_run_notes(wb, stats)

        # Sheet 10: Pipeline Fit Notes
        pipeline_fits = [o for o in opportunities
                        if o.tsm_pipeline_fit and "not accessible" not in o.tsm_pipeline_fit.lower()
                        and "no direct" not in o.tsm_pipeline_fit.lower()]
        self._write_pipeline_sheet(wb, pipeline_fits)

        # Remove default empty sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filepath)
        logger.info(f"Excel workbook saved: {filepath}")
        return str(filepath)

    def _write_master_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Write the main Master Opportunities sheet."""
        ws = wb.active
        ws.title = "Master Opportunities"

        # Headers
        for col_idx, (field, header, width) in enumerate(MASTER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 55)

        # Data rows
        for row_idx, opp in enumerate(opportunities, 2):
            opp_dict = opp.to_dict()
            for col_idx, (field, _, _) in enumerate(MASTER_COLUMNS, 1):
                value = opp_dict.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Add hyperlinks for URL columns
            self._add_hyperlink(ws, row_idx, 6, opp.listing_url)  # Listing URL
            self._add_hyperlink(ws, row_idx, 7, opp.application_url)  # Application URL

            # Color-code by overall score
            score = opp.overall_relevance_score
            if score >= 60:
                for col_idx in range(1, len(MASTER_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = HIGH_PRIORITY_FILL
            elif score >= 35:
                for col_idx in range(1, len(MASTER_COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = MEDIUM_PRIORITY_FILL

        # Freeze header row + auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Sort by overall score descending
        # (Data is pre-sorted in pipeline)

    def _write_filtered_sheet(self, wb: Workbook, sheet_name: str,
                               opportunities: List[Opportunity]):
        """Write a filtered subset sheet with key columns."""
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

        key_columns = [
            ("title", "Title", 40),
            ("opportunity_type", "Type", 12),
            ("funder", "Funder", 22),
            ("listing_url", "URL", 30),
            ("synopsis", "Synopsis", 45),
            ("tsm_fit_score", "TSM", 8),
            ("gmu_center_fit_score", "GMU", 8),
            ("overall_relevance_score", "Score", 8),
            ("deadline", "Deadline", 14),
            ("typical_award", "Award", 14),
            ("suggested_proposal_angle", "Proposal Angle", 40),
            ("why_this_could_work", "Why This Works", 50),
            ("final_recommendation", "Rec", 12),
            ("urgency", "Urgency", 12),
            ("recommended_next_step", "Next Step", 30),
        ]

        # Headers
        for col_idx, (_, header, width) in enumerate(key_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 55)

        # Data
        for row_idx, opp in enumerate(opportunities, 2):
            opp_dict = opp.to_dict()
            for col_idx, (field, _, _) in enumerate(key_columns, 1):
                value = opp_dict.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Hyperlink on URL column
            self._add_hyperlink(ws, row_idx, 4, opp.listing_url)

        ws.freeze_panes = "A2"
        if ws.dimensions:
            ws.auto_filter.ref = ws.dimensions

    def _write_award_analysis_sheet(self, wb: Workbook):
        """Write funder intelligence / past award analysis sheet."""
        from ..enrichment.award_analyzer import FUNDER_PROFILES

        ws = wb.create_sheet(title="Past Award Analysis")
        headers = ["Funder", "Typical Award", "Project Length",
                   "Preferences", "Favored Language", "Common Winners"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 40

        for row_idx, (name, profile) in enumerate(FUNDER_PROFILES.items(), 2):
            ws.cell(row=row_idx, column=1, value=name.title())
            ws.cell(row=row_idx, column=2, value=profile.get("typical_award", ""))
            ws.cell(row=row_idx, column=3, value=profile.get("project_length", ""))
            cell = ws.cell(row=row_idx, column=4, value=profile.get("preferences", ""))
            cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=5, value=profile.get("favored_language", ""))
            ws.cell(row=row_idx, column=6, value=profile.get("common_winners", ""))

        ws.freeze_panes = "A2"

    def _write_source_log(self, wb: Workbook, stats: RunStats):
        """Write source query log sheet."""
        ws = wb.create_sheet(title="Source Log")
        headers = ["Source", "Status", "Queries", "Results", "Errors"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 40

        ws.cell(row=2, column=1, value="Total Sources Queried")
        ws.cell(row=2, column=3, value=stats.sources_queried)
        ws.cell(row=3, column=1, value="Successful")
        ws.cell(row=3, column=2, value="OK")
        ws.cell(row=3, column=3, value=stats.sources_successful)
        ws.cell(row=4, column=1, value="Failed/Skipped")
        ws.cell(row=4, column=2, value="FAIL")
        ws.cell(row=4, column=3, value=stats.sources_failed)

        if stats.errors:
            for idx, err in enumerate(stats.errors[:20], 6):
                ws.cell(row=idx, column=1, value="Error")
                ws.cell(row=idx, column=5, value=str(err)[:200])

        ws.freeze_panes = "A2"

    def _write_run_notes(self, wb: Workbook, stats: RunStats):
        """Write run notes and summary sheet."""
        ws = wb.create_sheet(title="Run Notes")

        notes = [
            ("Agent Geronimo Run Summary", ""),
            ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Total Raw Opportunities", stats.total_raw),
            ("Total After Dedup", stats.total_deduped),
            ("High Priority TSM", stats.high_priority_tsm),
            ("High Priority GMU", stats.high_priority_gmu),
            ("Federal Opportunities", stats.federal_count),
            ("Foundation Opportunities", stats.foundation_count),
            ("Sources Queried", stats.sources_queried),
            ("Sources Successful", stats.sources_successful),
            ("Errors", len(stats.errors)),
            ("", ""),
            ("Notes", ""),
            ("- Scores are keyword-based relevance estimates (0-100)", ""),
            ("- Confidence reflects data completeness, not accuracy", ""),
            ("- All URLs should be independently verified before applying", ""),
            ("- Deadlines may change; always check the original listing", ""),
            ("- 'Why This Could Work' column synthesizes TSM + funder fit", ""),
            ("- Past Award Analysis sheet has pre-seeded funder intelligence", ""),
        ]

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 50

        for row_idx, (label, value) in enumerate(notes, 1):
            cell_a = ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=str(value))
            if row_idx == 1:
                cell_a.font = Font(bold=True, size=14)
            elif label and not label.startswith("-"):
                cell_a.font = Font(bold=True)

    def _write_pipeline_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Write TSM pipeline fit notes sheet."""
        ws = wb.create_sheet(title="Pipeline Fit Notes")
        headers = ["Title", "Funder", "TSM Fit", "Pipeline Fit",
                   "Explanation", "Concept Match", "Internal Files", "Why It Works"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        widths = [35, 20, 8, 20, 40, 20, 20, 50]
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        for row_idx, opp in enumerate(opportunities, 2):
            ws.cell(row=row_idx, column=1, value=opp.title)
            ws.cell(row=row_idx, column=2, value=opp.funder)
            ws.cell(row=row_idx, column=3, value=opp.tsm_fit_score)
            ws.cell(row=row_idx, column=4, value=opp.tsm_pipeline_fit)
            cell = ws.cell(row=row_idx, column=5, value=opp.tsm_pipeline_explanation)
            cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row_idx, column=6, value=opp.existing_concept_match)
            ws.cell(row=row_idx, column=7, value=opp.internal_file_references)
            cell = ws.cell(row=row_idx, column=8, value=opp.why_this_could_work)
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A2"

    @staticmethod
    def _add_hyperlink(ws, row: int, col: int, url: str):
        """Add a clickable hyperlink to a cell."""
        if url and url.startswith("http"):
            cell = ws.cell(row=row, column=col)
            try:
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
                # Shorten display text
                cell.value = url[:80] + "..." if len(url) > 80 else url
            except Exception:
                cell.value = url
